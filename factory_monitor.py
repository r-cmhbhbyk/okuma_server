# VERSION: V14
"""
Factory Machine Monitor
=======================
1. Clear old CSV files
2. Download fresh CSV from web interface
3. Analyze last hour:
   - Program cycle times + averages
   - Downtime detection with reasons
   - Timeline chart per machine
   - 7-day history (SQLite)
   - Telegram alert if idle > 60 min
4. Open HTML report in browser
"""

import os
import html as _html
import time
import sqlite3
import csv
import sys
import json
import subprocess
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from collections import defaultdict
from functools import lru_cache

# Selenium видалено — використовується Connect Plan WebAPI

# ── Configuration ─────────────────────────────────────────────────────────────
DOWNLOAD_DIR        = r"C:\Connectplan_raports"
OUTPUT_HTML         = os.path.join(DOWNLOAD_DIR, "index.html")
DB_FILE             = os.path.join(DOWNLOAD_DIR, "history.db")
LOG_FILE            = os.path.join(DOWNLOAD_DIR, "factory_monitor.log")
HOURS_BACK          = 24

# ── Connect Plan WebAPI ───────────────────────────────────────────────────────
API_BASE  = "http://192.168.1.210/FactoryMonitorSuiteSVC"
# ProcResID → повна назва станку (з GetMachineList)
PROC_RES_MAP = {
    1:  "M1_M560R-V-e_0712-100198",
    7:  "M2_M560R-V-e-M5V01235",
    6:  "T1_L300E-M_PEA351",
    3:  "T2_ L300-MYW-e_MYW197",
    8:  "T3_LB2000EXII_254633",
    12: "T4_LB3000EXII_247289",
}
ALERT_THRESHOLD_MIN = 45
S2S_GAP_THRESHOLD_MIN = 15  # Макс. розрив між циклами для start-to-start (хв); якщо більше — цикл не розтягується

# Повний список станків дільниці — використовується для графіків і розрахунку SITE
# Станки без даних отримують ефективність 0 і відображаються на графіку
ALL_MACHINES = ["M1_M560R-V-e_0712-100198", "M2_M560R-V-e-M5V01235",
                "T1_L300E-M_PEA351",        "T2_ L300-MYW-e_MYW197",
                "T3_LB2000EXII_254633"]

GITHUB_USER  = "wisefab1"
GITHUB_REPO  = "factory_monitor"
GITHUB_URL   = "https://wisefab1.github.io/factory_monitor/"

# ── Secrets (завантажуються з файлу, не зберігаються в коді) ─────────────────
_SECRETS_FILE = os.path.join(DOWNLOAD_DIR, "secrets.json")
def _load_secrets():
    try:
        with open(_SECRETS_FILE, encoding="utf-8") as _f:
            return json.load(_f)
    except Exception:
        return {}
_secrets         = _load_secrets()
TELEGRAM_TOKEN   = _secrets.get("telegram_token",   "")
TELEGRAM_CHAT_ID = _secrets.get("telegram_chat_id", "")
GITHUB_TOKEN     = _secrets.get("github_token",     "")

TARGET_TIME_FILE = r"\\wisefile\Wisefile\WF_Tootmine\Planeerimine\CNC toodete ajad pildid\CNC tehno.xlsm"
TARGET_CACHE_FILE = os.path.join(DOWNLOAD_DIR, "target_times_cache.json")  # Кеш файл

# ── Logging ───────────────────────────────────────────────────────────────────
def log(message: str):
    """Виводить повідомлення в консоль та зберігає у файл логів."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_line = f"[{timestamp}] {message}"
    print(log_line)
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(log_line + "\n")
    except Exception as e:
        print(f"Failed to write log: {e}")

# ── Excel Target Time ─────────────────────────────────────────────────────────
@lru_cache(maxsize=512)
def normalize_program_name(name: str) -> str:
    """Нормалізує назву програми для порівняння
    
    Порядок обробки:
    1. Видалення розширення
    2. Визначення операції та видалення (op1-5, p1-5, -1 до -5)
    3. Видалення останньої L або R
    4. Переведення у верхній регістр
    5. Видалення пробілів, дефісів, підкреслень
    6. Заміна 101 на 100
    
    Приклад: "WF861-100L-P2.MIN" → "WF861100"
    Приклад: "WF080-920-2.MIN" → "WF080920"
    """
    if not name:
        return ""
    
    # 1. Видаляємо розширення
    if '.' in name:
        name = name.rsplit('.', 1)[0]
    
    # 2. Визначаємо та видаляємо номер операції з кінця
    # Спочатку шукаємо op1-5 або p1-5
    name_lower = name.lower()
    found_op = False
    for op in ['op5', 'p5', 'op4', 'p4', 'op3', 'p3', 'op2', 'p2', 'op1', 'p1']:
        for variant in [f'-{op}', f'_{op}', op]:
            if name_lower.endswith(variant):
                name = name[:len(name) - len(variant)]
                found_op = True
                break
        if found_op:
            break
    
    # Якщо не знайшли op/p, шукаємо просто -1, -2, -3, -4, -5 в кінці
    if not found_op:
        for digit in ['5', '4', '3', '2', '1']:
            if name.endswith(f'-{digit}') or name.endswith(f'_{digit}'):
                name = name[:-2]
                break
    
    # 3. Видаляємо всі букви після останньої цифри
    # Знаходимо позицію останньої цифри
    last_digit_pos = -1
    for i in range(len(name) - 1, -1, -1):
        if name[i].isdigit():
            last_digit_pos = i
            break
    
    # Якщо знайшли цифру - обрізаємо все після неї
    if last_digit_pos >= 0 and last_digit_pos < len(name) - 1:
        name = name[:last_digit_pos + 1]
    
    # 4. Переводимо в верхній регістр
    name = name.upper()
    
    # 5. Видаляємо пробіли, дефіси, підкреслення
    name = name.replace(' ', '').replace('-', '').replace('_', '')
    
    # 6. Замінюємо 101 на 100 (WF861-100L та WF861_101L - одна деталь)
    name = name.replace('101', '100')
    
    return name

def parse_program_name(program_name: str) -> tuple:
    """Розбирає назву програми на базову назву та номер операції
    
    Args:
        program_name: повна назва програми (напр. "WF861-100L-P3.MIN")
    
    Returns:
        tuple: (normalized_base, operation_number)
        Приклад: "WF861-100L-P3.MIN" → ("WF861100", 3)
                 "WF080-920-2.MIN" → ("WF080920", 2)
    """
    normalized = normalize_program_name(program_name)  # "WF861100"
    operation = get_operation_number(program_name)      # 3
    return (normalized, operation)

_OP_SUFFIXES = [
    ('op5',5),('p5',5),('-p5',5),('-5',5),('_5',5),
    ('op4',4),('p4',4),('-p4',4),('-4',4),('_4',4),
    ('op3',3),('p3',3),('-p3',3),('-3',3),('_3',3),
    ('op2',2),('p2',2),('-p2',2),('-2',2),('_2',2),
    ('op1',1),('p1',1),('-p1',1),('-1',1),('_1',1),
]

@lru_cache(maxsize=512)
def get_operation_number(program_name: str) -> int:
    """Визначає номер операції з назви програми (OP1–OP5)."""
    if not program_name or program_name == "—":
        return 1
    base_lower = (program_name.rsplit('.', 1)[0] if '.' in program_name else program_name).lower()
    for suffix, op in _OP_SUFFIXES:
        if base_lower.endswith(suffix):
            return op
    return 1

def load_target_times():
    """Завантажує Target Time з Excel файлу з кешуванням
    
    Логіка:
    1. Перевіряємо чи доступний Excel файл
    2. Якщо доступний - порівнюємо timestamp з кешем
    3. Якщо Excel новіший - завантажуємо і оновлюємо кеш
    4. Якщо Excel не доступний - використовуємо кеш
    5. Якщо немає ні Excel ні кешу - повертаємо {}
    """
    excel_available = False
    excel_mtime = None
    cache_mtime = None

    # Перевіряємо доступність Excel файлу
    try:
        if os.path.exists(TARGET_TIME_FILE):
            excel_mtime = os.path.getmtime(TARGET_TIME_FILE)
            excel_available = True
            log(f"Excel file found: {TARGET_TIME_FILE}")
            log(f"Excel modified: {datetime.fromtimestamp(excel_mtime).strftime('%Y-%m-%d %H:%M:%S')}")
        else:
            log(f"Excel file not accessible: {TARGET_TIME_FILE}")
    except Exception as e:
        log(f"Cannot access Excel file: {e}")
        excel_available = False

    # Перевіряємо наявність кешу
    cache_exists = os.path.exists(TARGET_CACHE_FILE)
    if cache_exists:
        cache_mtime = os.path.getmtime(TARGET_CACHE_FILE)
        log(f"Cache file found: {TARGET_CACHE_FILE}")
        log(f"Cache modified: {datetime.fromtimestamp(cache_mtime).strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Вирішуємо звідки завантажувати
    use_excel = False
    
    if excel_available:
        # Завжди читаємо з Excel якщо він доступний — mtime на мережевому диску
        # може не оновлюватись одразу після збереження файлу
        log("✓ Excel accessible - will load from Excel")
        use_excel = True
    else:
        if cache_exists:
            # Excel недоступний, є кеш - використовуємо кеш
            log("⚠ Excel not accessible - will use cache")
            use_excel = False
        else:
            # Немає ні Excel ні кешу
            log("✗ No Excel and no cache - returning empty")
            return {}
    
    # Завантажуємо з Excel
    if use_excel:
        target_times = _load_from_excel()
        if target_times:
            # Зберігаємо в кеш
            _save_to_cache(target_times)
        return target_times
    
    # Завантажуємо з кешу
    else:
        return _load_from_cache()


def _load_from_excel():
    """Завантажує дані з Excel файлу
    
    Returns:
        dict: {(program, operation, machine): time} або {}
    """
    try:
        import openpyxl
        log(f"Loading target times from Excel...")
        
        # Спроба 1: Прямий доступ до UNC шляху
        try:
            wb = openpyxl.load_workbook(TARGET_TIME_FILE, read_only=True, data_only=True)
            log("✓ Direct UNC access successful")
        except Exception as e1:
            log(f"Direct UNC access failed: {e1}")
            
            # Спроба 2: Через pathlib
            try:
                from pathlib import Path
                path = Path(TARGET_TIME_FILE)
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                log("✓ Pathlib access successful")
            except Exception as e2:
                log(f"Pathlib access failed: {e2}")
                
                # Спроба 3: Через os.path.normpath
                try:
                    normalized = os.path.normpath(TARGET_TIME_FILE)
                    wb = openpyxl.load_workbook(normalized, read_only=True, data_only=True)
                    log("✓ Normalized path access successful")
                except Exception as e3:
                    log(f"All access methods failed: {e3}")
                    return {}
        
        # Шукаємо вкладку "Tehnoloogiad"
        if "Tehnoloogiad" not in wb.sheetnames:
            log(f"Warning: Sheet 'Tehnoloogiad' not found. Available: {wb.sheetnames}")
            wb.close()
            return {}
        
        ws = wb["Tehnoloogiad"]
        target_times = {}
        
        # Маппінг колонок: (станок_col, час_col)
        op_columns = {
            1: (11, 14),   # K, N
            2: (16, 19),   # P, S
            3: (21, 24),   # U, X
            4: (26, 29),   # Z, AC
            5: (31, 34),   # AE, AH
        }
        
        rows_processed = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            program_cell = row[0]
            if not program_cell or not program_cell.value:
                continue
            
            program_name = str(program_cell.value).strip()
            if not program_name:
                continue
            
            rows_processed += 1
            
            for op_num, (machine_col, time_col) in op_columns.items():
                machine_cell = row[machine_col - 1]
                time_cell = row[time_col - 1]
                
                if machine_cell and machine_cell.value and time_cell and time_cell.value:
                    machine = str(machine_cell.value).strip()
                    try:
                        time_val = float(time_cell.value)
                        key = (program_name, op_num, machine)
                        target_times[key] = time_val
                    except (ValueError, TypeError):
                        continue
        
        wb.close()
        log(f"✓ Loaded {len(target_times)} target times from {rows_processed} rows")
        return target_times
        
    except ImportError:
        log("Warning: openpyxl not installed")
        return {}
    except Exception as e:
        log(f"Error loading from Excel: {e}")
        return {}


def _save_to_cache(target_times):
    """Зберігає target times в JSON кеш
    
    Args:
        target_times: словник {(program, op, machine): time}
    """
    try:
        # Конвертуємо ключі-tuple в строки для JSON
        cache_data = {
            "updated": datetime.now().isoformat(),
            "count": len(target_times),
            "data": {
                f"{prog}|{op}|{machine}": time_val
                for (prog, op, machine), time_val in target_times.items()
            }
        }
        
        with open(TARGET_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, indent=2)
        
        log(f"✓ Cache saved: {len(target_times)} records")
        
    except Exception as e:
        log(f"Warning: Failed to save cache: {e}")


def _load_from_cache():
    """Завантажує target times з JSON кешу
    
    Returns:
        dict: {(program, operation, machine): time} або {}
    """
    try:
        with open(TARGET_CACHE_FILE, "r", encoding="utf-8") as f:
            cache_data = json.load(f)
        
        updated = cache_data.get("updated", "unknown")
        count = cache_data.get("count", 0)
        data = cache_data.get("data", {})
        
        log(f"✓ Loaded from cache: {count} records (updated: {updated})")
        
        # Конвертуємо строкові ключі назад в tuple
        target_times = {}
        for key_str, time_val in data.items():
            parts = key_str.split("|")
            if len(parts) == 3:
                prog, op_str, machine = parts
                try:
                    op = int(op_str)
                    target_times[(prog, op, machine)] = time_val
                except ValueError:
                    continue
        
        return target_times
        
    except Exception as e:
        log(f"Error loading from cache: {e}")
        return {}


def calculate_real_cycle_time(durations):
    """KDE, bandwidth=0.3. Знаходить найщільніший пік → середнє кластеру."""
    if not durations:
        return None
    if isinstance(durations[0], (tuple, list)):
        vals = [d for d, _ in durations if d > 0]
    else:
        vals = [d for d in durations if d > 0]
    if not vals:
        return None

    bandwidth = 0.3
    step = 0.05
    best_center, best_count = None, 0
    v = min(vals)
    while v <= max(vals) + step:
        count = sum(1 for x in vals if abs(x - v) <= bandwidth)
        if count > best_count:
            best_count, best_center = count, v
        v = round(v + step, 4)

    cluster = [x for x in vals if abs(x - best_center) <= bandwidth]
    return round(sum(cluster) / len(cluster), 2) if cluster else round(sum(vals) / len(vals), 2)


# ── Shared helpers ────────────────────────────────────────────────────────────
def _parse_ts(s: str) -> datetime:
    """Парсить рядок дати з CSV формату "%Y.%m.%d %H:%M:%S"."""
    return datetime.strptime(s, "%Y.%m.%d %H:%M:%S")

def _canonical_machine(name: str) -> str:
    """Повертає канонічну назву станку за першими двома символами.
    Приклад: "M1_..." → "M1_M560R-V-e_0712-100198" (з ALL_MACHINES).
    Якщо не знайдено в ALL_MACHINES — повертає оригінал.
    """
    if not name:
        return name
    prefix = name[:2].upper()
    for canonical in ALL_MACHINES:
        if canonical[:2].upper() == prefix:
            return canonical
    return name

def _work_window_min(date_str: str) -> int:
    """Повертає кількість робочих хвилин для дня тижня.
    пн/пт: 07:00–19:00 = 720 хв; вт-чт: 06:30–00:30 = 1080 хв; сб/нд: 0.
    """
    try:
        wd = datetime.strptime(date_str, "%Y-%m-%d").weekday()
        if wd in (0, 4): return 720
        if wd in (1, 2, 3): return 1080
    except Exception:
        pass
    return 0

# =============================================================================
# PART 1 — DOWNLOAD
# =============================================================================

def _api_get(endpoint: str, params: dict) -> list:
    """Виконує GET запит до Connect Plan WebAPI, повертає data[]."""
    qs = urllib.parse.urlencode(params)
    url = f"{API_BASE}/{endpoint}?{qs}"
    try:
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, timeout=15) as r:
            body = json.loads(r.read().decode("utf-8"))
        code = body.get("d", {}).get("code", -1)
        if str(code) != "0":
            log(f"✗ API error {code}: {body.get('d', {}).get('message')}")
            return []
        return body["d"]["data"]
    except Exception as e:
        log(f"✗ API request failed ({endpoint}): {e}")
        return []


def _api_ts(s: str) -> datetime:
    """Парсить дату з WebAPI формату '2026.04.18 21:54:44'."""
    return datetime.strptime(s, "%Y.%m.%d %H:%M:%S")


def fetch_from_api() -> tuple[list[dict], list[dict]]:
    """Замінює download_both_files() + load_csv() + завантаження mr_data.

    Повертає (rows, mr_data) — рядки в тому самому форматі що раніше давав CSV,
    сумісні з усім подальшим кодом аналізу.

    operation_history → GetOperationResult  (всі події RunState/Alarm/тощо)
    machining_results → GetMachiningResult  (цикли з Counter)
    """
    log("============================================================")
    log("FETCHING DATA FROM CONNECT PLAN WebAPI")
    log("============================================================")

    ids = ",".join(str(i) for i in PROC_RES_MAP.keys())
    now = datetime.now()
    start_dt = now.replace(hour=0, minute=0, second=0, microsecond=0)
    start_s  = start_dt.strftime("%Y/%m/%d %H:%M:%S")
    end_s    = now.strftime("%Y/%m/%d %H:%M:%S")

    # ── 1. GetOperationResult → rows ─────────────────────────────────
    log(f"── Fetching OperationResult {start_s} … {end_s} ──")
    raw = _api_get("v3/GetOperationResult", {
        "Specify":   "PROCRES",
        "ID":        ids,
        "StartDate": start_s,
        "EndDate":   end_s,
        "Sort":      0,
    })
    log(f"  Received {len(raw)} records")

    rows = []
    for r in raw:
        pid = r.get("ProcResID")
        mname = PROC_RES_MAP.get(pid, f"UNKNOWN_{pid}")
        try:
            ts = _api_ts(r["Date"])
        except Exception:
            continue
        rows.append({
            "_ts":              ts,
            "Date":             r["Date"],
            "MachineName":      mname,
            "RunState":         str(r.get("RunState", "0")),
            "ProgramFileName":  r.get("MainProgramFileName") or r.get("ProgramFileName") or "",
            "PowerOn":          str(r.get("PowerOn", "0")),
            "AlarmState":       str(r.get("AlarmState", "0")),
            "AlarmNo":          str(r.get("AlarmNo", "")),
            "AlarmMessage":     r.get("AlarmMessage") or r.get("AlarmString") or "",
            "LimitState":       str(r.get("LimitState", "0")),
            "ProgramStopState": str(r.get("ProgramStopState", "0")),
            "FeedHoldState":    str(r.get("FeedHoldState", "0")),
            "STMState":         str(r.get("STMState", "0")),
            "SetUp":            str(r.get("SetUp", "0")),
            "NoOperator":       str(r.get("NoOperator", "0")),
            "Wait":             str(r.get("Wait", "0")),
            "Maintenance":      str(r.get("Maintenance", "0")),
        })

    log(f"  Parsed {len(rows)} rows for analysis")

    # ── 2. GetMachiningResult → mr_data ──────────────────────────────
    log(f"── Fetching MachiningResult ──")
    raw_mr = _api_get("v3/GetMachiningResult", {
        "Specify":   "PROCRES",
        "ID":        ids,
        "StartDate": start_s,
        "EndDate":   end_s,
        "Sort":      0,
    })
    log(f"  Received {len(raw_mr)} machining records")

    mr_data = []
    for r in raw_mr:
        pid = r.get("ProcResID")
        mname = PROC_RES_MAP.get(pid, f"UNKNOWN_{pid}")
        try:
            ts = _api_ts(r["Date"])
        except Exception:
            ts = None
        mr_data.append({
            "_ts":             ts,
            "Date":            r.get("Date", ""),
            "MachineName":     mname,
            "ProgramFileName": r.get("MainProgramFileName") or r.get("ProgramFileName") or "",
            "RunStateTime":    str(r.get("RunStateTime", 0)),
            "Counter":         str(r.get("WorkCountACount") or r.get("Counter") or 0),
        })

    log(f"  Parsed {len(mr_data)} machining records")
    log("============================================================")
    log("FETCH COMPLETE")
    log("============================================================")
    return rows, mr_data


# ── Telegram ──────────────────────────────────────────────────────────────────
def send_telegram(message: str):
    try:
        url  = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
        data = urllib.parse.urlencode({
            "chat_id":    TELEGRAM_CHAT_ID,
            "text":       message,
            "parse_mode": "HTML",
        }).encode()
        urllib.request.urlopen(urllib.request.Request(url, data=data), timeout=10)
        log("Telegram alert sent")
    except Exception as e:
        log(f"Telegram error: {e}")

# ── SQLite ────────────────────────────────────────────────────────────────────
def init_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS daily_summary (
            date TEXT, machine TEXT, run_min INTEGER, down_min INTEGER,
            total_min INTEGER, cycles INTEGER, avg_cycle REAL, efficiency REAL,
            PRIMARY KEY (date, machine)
        )""")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS downtime_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT, machine TEXT, start_time TEXT, end_time TEXT,
            duration INTEGER, reason TEXT
        )""")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS cycle_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT, machine TEXT, program TEXT, start_time TEXT, end_time TEXT,
            duration INTEGER
        )""")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS hourly_stats (
            date TEXT, machine TEXT, hour INTEGER,
            run_min REAL, total_min REAL,
            PRIMARY KEY (date, machine, hour)
        )""")
    conn.commit()
    return conn


def save_to_db(conn, date_str, cycles, downtimes):
    for mname in cycles:
        c_list    = cycles[mname]
        d_data    = downtimes[mname]
        run_min   = d_data.get("total_run_all", d_data["total_run"])
        down_min  = d_data["total_down"]
        total_min = _work_window_min(date_str) or d_data.get("total_min", 0)
        run_min_working = d_data["total_run"]
        eff       = round(run_min_working / total_min * 100, 1) if total_min else 0
        avg_cycle = round(sum(c["duration"] for c in c_list) / len(c_list), 1) if c_list else 0
        conn.execute("""
            INSERT OR REPLACE INTO daily_summary
            (date,machine,run_min,down_min,total_min,cycles,avg_cycle,efficiency)
            VALUES (?,?,?,?,?,?,?,?)
        """, (date_str, mname, run_min, down_min, total_min, len(c_list), avg_cycle, eff))
        conn.execute("DELETE FROM cycle_events WHERE date=? AND machine=?", (date_str, mname))
        for c in c_list:
            conn.execute("""
                INSERT INTO cycle_events
                (date,machine,program,start_time,end_time,duration)
                VALUES (?,?,?,?,?,?)
            """, (date_str, mname, c.get("program", "—"),
                  c["start"].strftime("%H:%M") if c.get("start") else "—",
                  c["end"].strftime("%H:%M") if c.get("end") else "—",
                  c["duration"]))
        conn.execute("DELETE FROM downtime_events WHERE date=? AND machine=?", (date_str, mname))
        for d in d_data["downtimes"]:
            conn.execute("""
                INSERT INTO downtime_events
                (date,machine,start_time,end_time,duration,reason)
                VALUES (?,?,?,?,?,?)
            """, (date_str, mname,
                  d["start"].strftime("%H:%M"),
                  d["end"].strftime("%H:%M") if d.get("end") else "ongoing",
                  d["duration"], d["reason"]))

    # Hourly stats — розподіл run/total по годинах для кожної машини.
    # Використовується Today-графіком для відображення 7-денної історії.
    hr_run   = defaultdict(lambda: defaultdict(float))
    hr_total = defaultdict(lambda: defaultdict(float))
    def _spread(mname, start_dt, end_dt, is_run):
        cur = start_dt
        while cur < end_dt:
            hr_end = cur.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
            seg_min = (min(end_dt, hr_end) - cur).total_seconds() / 60
            if is_run:
                hr_run[mname][cur.hour] += seg_min
            hr_total[mname][cur.hour] += seg_min
            cur = hr_end
    for mname, c_list in cycles.items():
        for c in c_list:
            if c.get("start") and c.get("duration"):
                c_end = c.get("end") or (c["start"] + timedelta(minutes=c["duration"]))
                _spread(mname, c["start"], c_end, True)
    for mname, d_data in downtimes.items():
        for d in d_data.get("downtimes", []):
            if d.get("start") and d.get("duration"):
                d_end = d.get("end") or (d["start"] + timedelta(minutes=d["duration"]))
                _spread(mname, d["start"], d_end, False)
    conn.execute("DELETE FROM hourly_stats WHERE date=?", (date_str,))
    for mname, hr_map in hr_total.items():
        for h, total in hr_map.items():
            run = hr_run[mname].get(h, 0)
            conn.execute(
                "INSERT OR REPLACE INTO hourly_stats "
                "(date,machine,hour,run_min,total_min) VALUES (?,?,?,?,?)",
                (date_str, mname, h, round(run, 2), round(total, 2))
            )

    conn.commit()


def load_history(conn, machine: str, days: int = 7) -> list:
    cur = conn.execute("""
        SELECT date, efficiency, run_min, down_min, cycles, avg_cycle
        FROM daily_summary WHERE machine=? ORDER BY date DESC LIMIT ?
    """, (machine, days))
    return list(reversed(cur.fetchall()))


# ── Data processing ───────────────────────────────────────────────────────────
def get_counter_markers(mr_data, cycles_dict):
    """Повертає {machine_name: [datetime, ...]} — моменти COUNTER.MIN що реально
    потрапляють в межі циклу програми на цій машині (Counter >= 1).
    Відображаються як фіолетові лінії на таймлайні.
    """
    markers = defaultdict(list)
    for mr in mr_data:
        prog = mr.get("ProgramFileName", "")
        if not prog.upper().startswith("COUNTER"):
            continue
        try:
            ctr = int(mr.get("Counter", 0))
        except (ValueError, TypeError):
            ctr = 0
        if ctr < 1:
            continue
        machine = mr.get("MachineName", "")
        if not machine:
            continue
        marker_dt = mr.get("_ts")
        if marker_dt is None:
            continue
        # Додаємо мітку тільки якщо вона потрапляє в межі якогось циклу цієї машини
        for c in cycles_dict.get(machine, []):
            cs = c.get("start")
            ce = c.get("end")
            if cs and ce and cs <= marker_dt <= ce:
                markers[machine].append(marker_dt)
                break
    return markers


def split_cycles_by_counter(cycles_dict, counter_markers):
    """Розбиває цикли на підцикли по мітках COUNTER.MIN.

    COUNTER.MIN в machining_results записується в момент кінця циклу
    (Date = кінець). Тому мітка може збігатися з c_end або бути трохи пізніше.

    Мітка є роздільником якщо потрапляє у вікно [c_start+30s ... c_end+60s].
    Якщо мітка > c_end - використовуємо c_end як точку розрізу.
    Ongoing цикли не розбиваємо.
    """
    TOL_AFTER = timedelta(seconds=60)
    TOL_MIN   = timedelta(seconds=30)

    result = {}
    for mname, cycles in cycles_dict.items():
        markers = sorted(counter_markers.get(mname, []))
        new_cycles = []
        for c in cycles:
            c_start = c.get("start")
            c_end   = c.get("end")
            if not c_start or not c_end:
                new_cycles.append(c)
                continue
            inner = []
            for m in markers:
                if m <= c_start + TOL_MIN:
                    continue
                if m > c_end + TOL_AFTER:
                    continue
                cut = min(m, c_end)
                inner.append(cut)
            if not inner:
                new_cycles.append(c)
                continue
            inner = sorted(set(inner))
            if inner and inner[-1] == c_end:
                inner = inner[:-1]
            if not inner:
                new_cycles.append(c)
                continue
            boundaries = [c_start] + inner + [c_end]
            for i in range(len(boundaries) - 1):
                seg_start = boundaries[i]
                seg_end   = boundaries[i + 1]
                new_cycles.append({
                    "start":    seg_start,
                    "end":      seg_end,
                    "program":  c["program"],
                    "duration": round((seg_end - seg_start).total_seconds() / 60, 2),
                    "ongoing":  False,
                })
        result[mname] = new_cycles
    return result

def apply_start_to_start_cycles(cycles_dict, counter_markers, mr_data=None):
    """Перераховує межі циклів на рівні окремої програми.

    Критерій: програма використовує COUNTER якщо в machining_results є
    запис COUNTER з Counter>=1 в проміжку [c_start, c_end] хоча б одного циклу
    цієї програми на цій машині.

    З COUNTER: цикл = зелений сектор між маркерами (вже розрізані split_cycles_by_counter).
               Маркери не змінюються.

    БЕЗ COUNTER: цикл = від старту поточного до старту наступного циклу
                 тієї ж програми (start-to-start).
                 Маркери генеруються на кожному старті нового циклу.
    """
    # Будуємо індекс COUNTER подій з mr_data: {mname: [datetime, ...]}
    counter_events = defaultdict(list)
    if mr_data:
        for r in mr_data:
            prog = r.get("ProgramFileName", "")
            if not prog.upper().startswith("COUNTER"):
                continue
            try:
                ctr = int(r.get("Counter", 0))
            except (ValueError, TypeError):
                ctr = 0
            if ctr < 1:
                continue
            mname = r.get("MachineName", "")
            ts = r.get("_ts")
            if mname and ts:
                counter_events[mname].append(ts)

    def prog_has_counter(mname, prog_cycles):
        """Перевіряє чи хоча б один цикл програми мав COUNTER всередині."""
        cevents = sorted(counter_events.get(mname, []))
        if not cevents:
            return False
        for c in prog_cycles:
            cs = c.get("start")
            ce = c.get("end")
            if not cs or not ce:
                continue
            for ct in cevents:
                if cs <= ct <= ce:
                    return True
        return False

    new_cycles = {}
    new_markers = dict(counter_markers)

    for mname, cycles in cycles_dict.items():
        # Групуємо цикли по програмах
        by_prog = defaultdict(list)
        for c in cycles:
            if c.get("start") and not c.get("program", "").upper().startswith("COUNTER"):
                by_prog[c["program"]].append(c)

        result = []
        extra_markers = list(counter_markers.get(mname, []))

        for prog, prog_cycles in by_prog.items():
            prog_cycles = sorted(prog_cycles, key=lambda c: c["start"])

            if prog_has_counter(mname, prog_cycles):
                # ── З COUNTER: цикли вже правильні (split_cycles_by_counter) ──
                for c in prog_cycles:
                    if "cycle_time" not in c:
                        c = dict(c, cycle_time=c.get("duration"))
                    result.append(c)
            else:
                # ── БЕЗ COUNTER: start-to-start ──────────────────────────────
                for i, c in enumerate(prog_cycles):
                    c_start = c["start"]
                    green_duration = c.get("duration", 0)  # реальний час зеленого сектору
                    if i + 1 < len(prog_cycles):
                        next_start = prog_cycles[i + 1]["start"]
                        s2s_candidate = (next_start - c_start).total_seconds() / 60
                        # Не розтягуємо цикл якщо розрив між циклами більше S2S_GAP_THRESHOLD_MIN
                        # — це означає що станок був вимкнений або стояв між циклами
                        if s2s_candidate > S2S_GAP_THRESHOLD_MIN:
                            c_end = c.get("end")
                        else:
                            c_end = next_start
                    else:
                        c_end = c.get("end")
                    s2s_duration = round((c_end - c_start).total_seconds() / 60, 2) if c_end else green_duration
                    # Маркер на старті ПЕРШОГО циклу програми (cycle = s2s першого)
                    if i == 0:
                        extra_markers.append((c_start, prog, s2s_duration))
                    # Маркер = старт наступного циклу, cycle = s2s наступного циклу
                    if i + 1 < len(prog_cycles):
                        next_c = prog_cycles[i + 1]
                        next_c_end = prog_cycles[i + 2]["start"] if i + 2 < len(prog_cycles) else next_c.get("end")
                        next_s2s = round((next_c_end - next_c["start"]).total_seconds() / 60, 2) if next_c_end else next_c.get("duration", 0)
                        extra_markers.append((c_end, prog, next_s2s))
                    result.append({
                        "start":      c_start,
                        "end":        c_end,
                        "program":    prog,
                        "duration":   green_duration,   # Duration колонка = зелений сектор
                        "cycle_time": s2s_duration,     # Cycle колонка = start-to-start
                        "ongoing":    c.get("ongoing", False) if i + 1 >= len(prog_cycles) else False,
                    })

        new_cycles[mname] = sorted(result, key=lambda c: c["start"])
        if extra_markers:
            # extra_markers містить (datetime, prog, green_dur) або просто datetime
            plain = []
            prog_markers = {}   # {datetime: prog}
            green_durs   = {}   # {datetime: green_dur}
            for m in extra_markers:
                if isinstance(m, tuple):
                    dt = m[0]; pr = m[1]
                    gd = m[2] if len(m) > 2 else None
                    plain.append(dt)
                    prog_markers[dt] = pr
                    if gd is not None:
                        green_durs[dt] = gd
                else:
                    plain.append(m)
            # Зберігаємо в new_markers як datetime
            existing = list(counter_markers.get(mname, []))
            new_markers[mname] = sorted(set(plain + existing))
            new_markers[f"__prog_{mname}"]  = prog_markers
            new_markers[f"__green_{mname}"] = green_durs

    return new_cycles, new_markers

def add_runstate_boundary_markers(counter_markers, rows, counter_machines):
    """Додає маркери на межах RunState 1↔0 і 0↔1 — тільки для машин що реально мають COUNTER.

    counter_machines — множина машин з get_counter_markers (до start-to-start розширення).
    """
    machines = defaultdict(list)
    for r in rows:
        machines[r["MachineName"]].append(r)

    result = dict(counter_markers)

    for mname, mrows in machines.items():
        if mname not in counter_machines:
            continue
        mrows = sorted(mrows, key=lambda r: r["Date"])
        existing = set(counter_markers.get(mname, []))

        prev_run = None
        for r in mrows:
            run = r["RunState"]
            if prev_run is not None and run != prev_run:
                existing.add(r["_ts"])
            prev_run = run

        result[mname] = sorted(existing)

    return result

def filter_last_hours(rows, hours):
    last_ts = max(r["_ts"] for r in rows)
    # Для 24 годин — починаємо з 00:00 того ж дня
    if hours >= 24:
        cutoff = last_ts.replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        cutoff = last_ts - timedelta(hours=hours)
    return [r for r in rows if r["_ts"] >= cutoff], cutoff, last_ts

def analyze_cycles(rows):
    machines = defaultdict(list)
    for r in rows:
        machines[r["MachineName"]].append(r)
    result = {}
    for mname, mrows in machines.items():
        mrows.sort(key=lambda r: r["Date"])
        cycles, prev_run, prev_prog_parsed, cycle_start, cycle_prog = [], None, None, None, ""
        for r in mrows:
            ts, run, prog = r["_ts"], r["RunState"], r["ProgramFileName"]
            prog_parsed = parse_program_name(prog)  # (base, operation)

            if run == "1":
                if prev_run in (None, "0"):
                    cycle_start, cycle_prog = ts, prog
                elif prev_run == "1" and prog_parsed != prev_prog_parsed and cycle_start:
                    cycles.append({"start": cycle_start, "end": ts, "program": cycle_prog,
                                    "duration": round((ts - cycle_start).total_seconds() / 60, 2)})
                    cycle_start, cycle_prog = ts, prog
            elif prev_run == "1" and run == "0" and cycle_start:
                cycles.append({"start": cycle_start, "end": ts, "program": cycle_prog,
                                "duration": round((ts - cycle_start).total_seconds() / 60, 2)})
                cycle_start = None

            prev_run = run
            prev_prog_parsed = prog_parsed

        if cycle_start:
            last_ts = mrows[-1]["_ts"]
            cycles.append({"start": cycle_start, "end": None, "program": cycle_prog,
                           "duration": round((last_ts - cycle_start).total_seconds() / 60, 2),
                           "ongoing": True})
        result[mname] = cycles
    return result

def _is_in_efficiency_window(ts, weekend_first, weekend_last):
    """Перевіряє чи timestamp потрапляє в робоче вікно дня тижня."""
    wd   = ts.weekday()
    hour = ts.hour + ts.minute / 60.0 + ts.second / 3600.0
    if wd in (2, 3, 4) and hour < 0.5: return True        # ніч 00:00–00:30 після Вт/Ср/Чт
    if wd in (0, 4):    return 7.0 <= hour < 19.0         # Пн, Пт: 07:00–19:00
    if wd in (1, 2, 3): return hour >= 6.5                # Вт-Чт: 06:30–00:00
    if wd in (5, 6):                                       # Сб, Нд
        if weekend_first is None or weekend_last is None:
            return False
        return weekend_first <= ts <= weekend_last
    return False

def analyze_downtime(rows):
    machines = defaultdict(list)
    for r in rows:
        machines[r["MachineName"]].append(r)
    result = {}
    for mname, mrows in machines.items():
        mrows.sort(key=lambda r: r["Date"])
        downtimes, prev_run, dt_start, dt_reason = [], None, None, ""

        # Знаходимо межі вихідного дня (перший/останній запис RunState=1)
        weekend_first = weekend_last = None
        for r in mrows:
            ts = r["_ts"]
            if ts.weekday() in (5, 6) and r["RunState"] == "1":
                if weekend_first is None:
                    weekend_first = ts
                weekend_last = ts

        filtered_rows = [
            r for r in mrows
            if _is_in_efficiency_window(r["_ts"], weekend_first, weekend_last)
        ]

        for r in mrows:
            ts, run = r["_ts"], r["RunState"]
            if run == "0":
                if   r["AlarmState"]       == "1": reason = "Alarm: " + (r["AlarmMessage"] or r["AlarmNo"] or "—")
                elif r["PowerOn"]          == "0": reason = "Power off"
                elif r["SetUp"]            == "1": reason = "Setup"
                elif r["Maintenance"]      == "1": reason = "Maintenance"
                elif r["NoOperator"]       == "1": reason = "No operator"
                elif r["Wait"]             == "1": reason = "Waiting"
                elif r["FeedHoldState"]    == "1": reason = "Feed Hold"
                elif r["ProgramStopState"] == "1": reason = "Program Stop"
                else:                              reason = "Idle"
            else:
                reason = ""
            if prev_run in (None, "1") and run == "0":
                dt_start, dt_reason = ts, reason
            elif prev_run == "0" and run == "1" and dt_start:
                dur = round((ts - dt_start).total_seconds() / 60, 2)
                if dur > 0:
                    downtimes.append({"start": dt_start, "end": ts,
                                      "duration": dur, "reason": dt_reason})
                dt_start = None
            prev_run = run
        if dt_start:
            last_ts = mrows[-1]["_ts"]
            dur = round((last_ts - dt_start).total_seconds() / 60, 2)
            if dur > 0:
                downtimes.append({"start": dt_start, "end": None, "duration": dur,
                                  "reason": dt_reason, "ongoing": True})

        result[mname] = {
            "downtimes":  downtimes,
            "total_run":  sum(1 for r in filtered_rows if r["RunState"] == "1"),
            "total_down": sum(1 for r in filtered_rows if r["RunState"] == "0"),
            "total_min":  len(filtered_rows),
            "total_run_all": sum(1 for r in mrows if r["RunState"] == "1"),
        }
    return result

def split_timeline_by_counter(timeline_data, counter_markers, period_from, period_to):
    """Розрізає зелені сегменти таймлайну по мітках COUNTER.MIN.

    COUNTER.MIN записується в кінці циклу, тому використовуємо вікно:
    [seg_start+30s ... seg_end+60s] - аналогічно до split_cycles_by_counter.
    """
    total_sec = max((period_to - period_from).total_seconds(), 1)
    TOL_AFTER = timedelta(seconds=60)
    TOL_MIN   = timedelta(seconds=30)

    def pct(dt):
        return (dt - period_from).total_seconds() / total_sec * 100

    def dt_from_pct(p):
        return period_from + timedelta(seconds=p / 100 * total_sec)

    result = {}
    for mname, segments in timeline_data.items():
        markers = sorted(counter_markers.get(mname, []))
        new_segs = []
        for seg in segments:
            if seg["state"] != "1" or not markers:
                new_segs.append(seg)
                continue
            seg_start_dt = dt_from_pct(seg["x"])
            seg_end_dt   = dt_from_pct(seg["x"] + seg["w"])
            inner = []
            for m in markers:
                if m <= seg_start_dt + TOL_MIN:
                    continue
                if m > seg_end_dt + TOL_AFTER:
                    continue
                cut = min(m, seg_end_dt)
                inner.append(cut)
            if not inner:
                new_segs.append(seg)
                continue
            inner = sorted(set(inner))
            if inner and inner[-1] == seg_end_dt:
                inner = inner[:-1]
            if not inner:
                new_segs.append(seg)
                continue
            boundaries_dt = [seg_start_dt] + inner + [seg_end_dt]
            base_id = seg["id"]
            for i in range(len(boundaries_dt) - 1):
                b_start = boundaries_dt[i]
                b_end   = boundaries_dt[i + 1]
                x = pct(b_start)
                w = pct(b_end) - x
                if w > 0.01:
                    new_segs.append({
                        "x":     x,
                        "w":     w,
                        "state": "1",
                        "label": seg["label"],
                        "start": b_start.strftime("%H:%M"),
                        "end":   b_end.strftime("%H:%M"),
                        "id":    f"{base_id}_{i}",
                    })
        result[mname] = new_segs
    return result


def build_timeline_data(rows, period_from, period_to):
    machines  = defaultdict(list)
    for r in rows:
        machines[r["MachineName"]].append(r)
    total_sec = max((period_to - period_from).total_seconds(), 1)
    result    = {}

    for mname, mrows in machines.items():
        mrows.sort(key=lambda r: r["Date"])
        segments  = []
        seg_start = period_from
        seg_state = None
        seg_label = ""
        seg_idx   = 0

        def _get_label(r):
            run = r["RunState"]
            if run == "1":
                return r["ProgramFileName"] or "Running"
            if   r["AlarmState"]       == "1": return "Alarm: " + (r["AlarmMessage"] or r["AlarmNo"] or "—")
            elif r["PowerOn"]          == "0": return "Power off"
            elif r["SetUp"]            == "1": return "Setup"
            elif r["Maintenance"]      == "1": return "Maintenance"
            elif r["NoOperator"]       == "1": return "No operator"
            elif r["Wait"]             == "1": return "Waiting"
            elif r["FeedHoldState"]    == "1": return "Feed Hold"
            elif r["ProgramStopState"] == "1": return "Program Stop"
            return "Idle"

        for r in mrows:
            if r["ProgramFileName"].upper().startswith("COUNTER"):
                continue  # службовий рядок — не розриваємо сегмент
            ts, run = r["_ts"], r["RunState"]
            lbl = _get_label(r)
            if seg_state is None:
                seg_state, seg_start, seg_label = run, ts, lbl
            elif run != seg_state or (run == "1" and lbl != seg_label):
                # нова програма або зміна стану — закриваємо сегмент
                x = (seg_start - period_from).total_seconds() / total_sec * 100
                w = (ts - seg_start).total_seconds() / total_sec * 100
                if w > 0.05:
                    segments.append({
                        "x": x, "w": w, "state": seg_state,
                        "label": seg_label,
                        "start": seg_start.strftime("%H:%M"),
                        "end":   ts.strftime("%H:%M"),
                        "id":    f"{mname.split('_')[0]}_{seg_idx}",
                    })
                    seg_idx += 1
                seg_start, seg_state, seg_label = ts, run, lbl

        if seg_state is not None:
            x = (seg_start - period_from).total_seconds() / total_sec * 100
            w = (period_to - seg_start).total_seconds() / total_sec * 100
            if w > 0.05:
                segments.append({
                    "x": x, "w": w, "state": seg_state,
                    "label": seg_label,
                    "start": seg_start.strftime("%H:%M"),
                    "end":   period_to.strftime("%H:%M"),
                    "id":    f"{mname.split('_')[0]}_{seg_idx}",
                })
        result[mname] = segments
    return result

# ── GitHub Pages publish ──────────────────────────────────────────────────────
def publish_to_github(html: str) -> bool:
    """Push index.html to GitHub Pages via API — no git install required."""
    import base64
    import traceback
    try:
        api     = f"https://api.github.com/repos/{GITHUB_USER}/{GITHUB_REPO}/contents/index.html"
        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Content-Type":  "application/json",
            "Accept":        "application/vnd.github+json",
        }
        log(f"GitHub API URL: {api}")
        log(f"GitHub User: {GITHUB_USER}")
        log(f"GitHub Repo: {GITHUB_REPO}")
        
        # Отримуємо SHA якщо файл вже існує
        sha = None
        try:
            log("Checking if index.html exists...")
            req = urllib.request.Request(api, headers=headers)
            with urllib.request.urlopen(req, timeout=10) as r:
                data = json.loads(r.read().decode())
                sha = data["sha"]
                log(f"File exists, SHA: {sha[:8]}...")
        except urllib.error.HTTPError as e:
            if e.code == 404:
                log("File doesn't exist, will create new")
            else:
                log(f"HTTP Error checking file: {e.code} {e.reason}")
                raise
        
        # Пушимо файл
        payload = {
            "message": f"update {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            "content": base64.b64encode(html.encode("utf-8")).decode(),
            "branch": "main"  # явно вказуємо гілку
        }
        if sha:
            payload["sha"] = sha
        
        log(f"Uploading HTML ({len(html)} bytes)...")
        req = urllib.request.Request(
            api, data=json.dumps(payload).encode(),
            headers=headers, method="PUT"
        )
        with urllib.request.urlopen(req, timeout=30) as r:
            response = json.loads(r.read().decode())
            log(f"Upload successful!")
            log(f"Commit SHA: {response['commit']['sha'][:8]}...")
        
        log(f"✓ Published: {GITHUB_URL}")
        return True
    except urllib.error.HTTPError as e:
        log(f"✗ GitHub HTTP Error: {e.code} {e.reason}")
        try:
            error_body = e.read().decode()
            log(f"Error details: {error_body}")
        except:
            pass
        return False
    except Exception as e:
        log(f"✗ GitHub publish error: {type(e).__name__}: {e}")
        import io
        s = io.StringIO()
        traceback.print_exc(file=s)
        log(s.getvalue())
        return False

def check_and_alert(downtimes, period_to, cycles, excel_targets):
    """Перевіряє простої та відправляє Telegram алерти

    Умови відправлення алерту:
    1. Є новий невідрапортований простій ≥45 хв, АБО
    2. Є старий ongoing простій що збільшився на +45 хв, АБО
    3. Є різниця між Calculated та Target >5%

    Додаткові правила:
    - Не відправляємо з 20:00 до 08:00
    - Закінчені простої не повторюємо
    - Всі повідомлення відправляємо не частіше ніж раз на годину
    - Через 24 години скидаємо список
    """
    log("── Step 3.6: Checking alerts (V14) ──")

    current_hour = datetime.now().hour
    current_time = datetime.now()

    log(f"Current hour: {current_hour}, downtimes: {len(downtimes)} machines, cycles: {len(cycles)} machines")

    # 1. Перевіряємо тихі години (20:00 - 08:00)
    if current_hour >= 20 or current_hour < 8:
        log("Silent hours (20:00-08:00) - no alerts sent")
        return

    # 2. Перевіряємо чи минула 1 година з останнього повідомлення (маркер-файл)
    telegram_marker = os.path.join(DOWNLOAD_DIR, "last_telegram_sent.txt")
    try:
        if os.path.exists(telegram_marker):
            marker_mtime = os.path.getmtime(telegram_marker)
            minutes_since_last = (time.time() - marker_mtime) / 60
            log(f"Telegram marker exists, last sent {minutes_since_last:.0f} min ago")
            if minutes_since_last < 55:
                log(f"Less than 55 min since last Telegram — SKIPPING entire alert check")
                return
            else:
                log(f"More than 55 min — will send if needed")
        else:
            log("No telegram marker file — first run, will send")
    except Exception as e:
        log(f"Telegram marker check error: {e} — will proceed")

    # Файл для збереження вже відправлених алертів
    sent_alerts_file = os.path.join(DOWNLOAD_DIR, "sent_alerts.json")
    
    # Завантажуємо список відправлених алертів
    sent_alerts = {}
    reset_needed = False
    last_reset_str = current_time.isoformat()

    if os.path.exists(sent_alerts_file):
        try:
            with open(sent_alerts_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                last_reset = data.get("last_reset", "")

                if last_reset:
                    last_reset_dt = datetime.fromisoformat(last_reset)
                    hours_since_reset = (current_time - last_reset_dt).total_seconds() / 3600
                    if hours_since_reset > 24:
                        reset_needed = True
                    else:
                        sent_alerts = data.get("alerts", {})
                        last_reset_str = last_reset  # зберігаємо оригінальний час скидання
                else:
                    reset_needed = True
        except:
            reset_needed = True
    else:
        reset_needed = True
    
    if reset_needed:
        sent_alerts = {}
    
    # Збираємо алерти про простої
    downtime_alerts = []
    
    log(f"Checking downtimes for {len(downtimes)} machines...")
    for mname, dd in downtimes.items():
        machine_downtimes = dd.get("downtimes", [])
        log(f"  {mname}: {len(machine_downtimes)} downtime events")
        
        for d in machine_downtimes:
            duration = d["duration"]
            is_ongoing = not d.get("end")
            log(f"    - Duration: {duration} min, Ongoing: {is_ongoing}, Threshold: {ALERT_THRESHOLD_MIN} min")
            
            if duration >= ALERT_THRESHOLD_MIN:
                # Унікальний ключ: машина + час початку (не змінюється для ongoing)
                alert_key = f"downtime_{mname}_{d['start'].strftime('%Y-%m-%d_%H:%M')}"
                already_sent = alert_key in sent_alerts
                
                log(f"    - Qualifies for alert! Key: {alert_key}, Already sent: {already_sent}")
                
                # Перевіряємо чи вже відправляли алерт для цього простою
                if already_sent:
                    # Якщо це ongoing простій - перевіряємо чи треба повторити
                    if is_ongoing:
                        prev_alert = sent_alerts[alert_key]
                        prev_duration = prev_alert.get("duration", 0)
                        additional_duration = duration - prev_duration
                        
                        log(f"    - Ongoing: prev_duration={prev_duration}, additional={additional_duration}")
                        
                        # Якщо протривав ще мінімум 45 хв - повторюємо алерт
                        if additional_duration >= 45:
                            log(f"    - ✓ Adding repeat alert (additional {additional_duration} min)")
                            downtime_alerts.append((mname, d, alert_key, True))  # True = repeat
                        else:
                            log(f"    - ✗ Not enough additional time ({additional_duration} < 45 min)")
                    else:
                        log(f"    - ✗ Downtime finished, not repeating")
                    # Якщо простій закінчився (є end) - НЕ повторюємо, пропускаємо
                else:
                    # Новий простій - відправляємо
                    log(f"    - ✓ Adding new alert")
                    downtime_alerts.append((mname, d, alert_key, False))  # False = new
            else:
                log(f"    - ✗ Below threshold ({duration} < {ALERT_THRESHOLD_MIN} min)")
    
    log(f"Found {len(downtime_alerts)} downtime alerts")
    
    # Збираємо алерти про перевищення Target >5%
    target_alerts = []
    on_target = []  # Програми в нормі (≤5%)
    no_norm_alerts = []  # Програми без норми в Excel
    machines_checked = set()  # Унікальні машини які перевірялись
    machines_with_issues = set()  # Машини з проблемами
    machines_no_norm = set()  # Машини де є цикли але нема норми
    
    log(f"Checking cycle times for {len(cycles)} machines...")
    excel_norm = [
        (normalize_program_name(p), op, normalize_program_name(m), t)
        for (p, op, m), t in excel_targets.items()
    ]
    # DEBUG: показуємо зразок машин з Excel
    unique_machines_excel = sorted(set(em for _, _, em, _ in excel_norm))
    log(f"  Excel machine norms (sample): {unique_machines_excel[:10]}")
    for mname, c_list in cycles.items():
        machine_short = mname.split("_")[0] if "_" in mname else mname
        machine_norm = normalize_program_name(machine_short)
        log(f"  Machine: {machine_short} → norm={machine_norm}")

        # Групуємо по програмах
        by_prog = {}
        for c in c_list:
            prog_name = c["program"] or "—"
            if prog_name not in by_prog:
                by_prog[prog_name] = []
            by_prog[prog_name].append(c)

        for prog, prog_cycles in by_prog.items():
            # Розраховуємо Calculated — повний час блоку (run + setup) = start-to-start
            sorted_pc = sorted(prog_cycles, key=lambda c: c["start"])
            block_times = []
            for idx_c, cc in enumerate(sorted_pc):
                if idx_c + 1 < len(sorted_pc) and cc.get("start") and sorted_pc[idx_c + 1].get("start"):
                    bt = round((sorted_pc[idx_c + 1]["start"] - cc["start"]).total_seconds() / 60, 2)
                    if bt > 0:
                        block_times.append(bt)
                else:
                    ct = cc.get("cycle_time") or cc.get("duration", 0)
                    if ct and ct > 0:
                        block_times.append(ct)
            if block_times:
                calc_target = calculate_real_cycle_time(block_times)
            else:
                calc_target = None

            if calc_target is None:
                continue

            # Визначаємо операцію
            op_num = get_operation_number(prog)
            prog_normalized = normalize_program_name(prog)

            # Шукаємо Excel Target
            excel_target = None
            prog_hits = [(ep, eop, em, t) for ep, eop, em, t in excel_norm if ep == prog_normalized]
            if not prog_hits:
                log(f"    {prog} (norm={prog_normalized}, machine={machine_norm}, op={op_num}): No prog match in Excel")
            else:
                log(f"    {prog} (norm={prog_normalized}, machine={machine_norm}, op={op_num}): {len(prog_hits)} prog matches, machines={[x[2] for x in prog_hits[:5]]}")
            for ep_norm, eop, em_norm, time_val in excel_norm:
                if prog_normalized == ep_norm and op_num == eop and machine_norm == em_norm:
                    excel_target = time_val
                    break

            # Якщо є Target
            if excel_target:
                machines_checked.add(machine_short)  # Додаємо до перевірених
                diff_pct = ((calc_target - excel_target) / excel_target) * 100
                log(f"    {prog}: Calculated={calc_target}, Target={excel_target}, Diff={round(diff_pct, 1)}%")

                if abs(diff_pct) > 5:
                    log(f"    ✓ Adding target alert (difference >5%)")
                    target_alerts.append((machine_short, prog, calc_target, excel_target, diff_pct, len(prog_cycles)))
                    machines_with_issues.add(machine_short)
                else:
                    log(f"    ✓ On target (difference ≤5%)")
                    on_target.append((machine_short, prog, calc_target, excel_target, diff_pct, len(prog_cycles)))
            else:
                log(f"    {prog}: No target found in Excel")
                no_norm_alerts.append((machine_short, prog, calc_target, len(prog_cycles)))
                machines_no_norm.add(machine_short)
    
    log(f"Found {len(target_alerts)} target alerts, {len(no_norm_alerts)} no-norm alerts")

    # Підраховуємо статистику
    total_machines = len(machines_checked | machines_no_norm)
    machines_with_issues_count = len(machines_with_issues)
    machines_no_norm_count = len(machines_no_norm - machines_with_issues)
    machines_ok_count = total_machines - machines_with_issues_count - machines_no_norm_count

    # Збираємо ВСІ поточні простої (ongoing) для включення в будь-яке повідомлення
    all_ongoing = []
    for mname, dd in downtimes.items():
        short = mname.split("_")[0] if "_" in mname else mname
        for d in dd.get("downtimes", []):
            if not d.get("end"):  # тільки ongoing
                all_ongoing.append((short, d))
    all_ongoing.sort(key=lambda x: x[1]["duration"], reverse=True)

    # Формуємо повідомлення
    if downtime_alerts or target_alerts or on_target or no_norm_alerts or all_ongoing:
        # Є проблеми або ongoing простої
        lines = [
            f"⚠️ <b>Factory Alert</b>  <i>V14</i>",
            f"📅 Data: {period_to.strftime('%H:%M')}  |  Sent: {current_time.strftime('%H:%M')}",
            f"🔗 <a href=\"{GITHUB_URL}\">Open report</a>\n"
        ]

        # Summary
        status_parts = []
        if machines_ok_count > 0:
            status_parts.append(f"✅ {machines_ok_count} OK")
        if machines_with_issues_count > 0:
            status_parts.append(f"🔴 {machines_with_issues_count} off target")
        if machines_no_norm_count > 0:
            status_parts.append(f"❓ {machines_no_norm_count} no norm")

        if total_machines > 0:
            lines.append(f"📊 <b>Status:</b> {' | '.join(status_parts)}")
            lines.append("")  # Порожня лінія
    else:
        # Все ОК - відправляємо позитивне повідомлення
        lines = [
            f"✅ <b>All Systems Normal</b>  <i>V14</i>",
            f"📅 Data: {period_to.strftime('%H:%M')}  |  Sent: {current_time.strftime('%H:%M')}",
            f"🔗 <a href=\"{GITHUB_URL}\">Open report</a>\n",
            f"📊 All {total_machines} machines within target cycle times"
        ]

    log(f"Preparing to send: {len(downtime_alerts)} downtime alerts, {len(target_alerts)} target alerts, {len(no_norm_alerts)} no-norm alerts, {len(all_ongoing)} ongoing")

    # ── Поточні простої (всі ongoing) ──────────────────────────────────
    if all_ongoing:
        lines.append("\n⏸ <b>Current downtime:</b>")
        for short, d in all_ongoing:
            dur = round(d['duration'])
            lines.append(
                f"  🔴 <b>{short}</b>  {d['start'].strftime('%H:%M')}–ongoing"
                f"  <b>{dur} min</b>  {_html.escape(d['reason'])}"
            )

    # Додаємо алерти про великі простої (>= threshold) з оновленням sent_alerts
    if downtime_alerts:
        lines.append("")
        for mname, d, alert_key, _ in downtime_alerts:
            # Оновлюємо інформацію про алерт
            sent_alerts[alert_key] = {
                "machine": mname,
                "start": d['start'].strftime('%Y-%m-%d %H:%M'),
                "duration": d['duration'],
                "last_alert": current_time.isoformat()
            }
    
    # Додаємо секцію Cycle Time (всі програми з нормою)
    all_cycle_entries = target_alerts + on_target
    if all_cycle_entries:
        lines.append("\n\n⚙️ <b>Cycle Time:</b>")
        for machine, prog, calc, target, diff_pct, n_cycles in all_cycle_entries:
            abs_diff = abs(diff_pct)
            if abs_diff <= 5:
                status = f'<b>{round(abs_diff, 1)}% {"slower" if diff_pct > 0 else "faster"}</b> 🟢'
            elif diff_pct > 0:
                status = f'<b>{round(abs_diff, 1)}% slower</b> 🔴'
            else:
                status = f'<b>{round(abs_diff, 1)}% faster</b> 🔵'

            lines.append(
                f"\n📊 <b>{machine}</b> - {prog}"
                f"\n   Calculated: {calc} min | Target: {target} min"
                f"\n   Cycles: {n_cycles}"
                f"\n   {status}"
            )

    # Додаємо алерти про відсутність норми
    if no_norm_alerts:
        lines.append("\n\n❓ <b>Norm missing:</b>")
        for machine, prog, calc, n_cycles in no_norm_alerts:
            lines.append(
                f"\n📋 <b>{machine}</b> - {prog}"
                f"\n   Cycle: {calc} min | Cycles: {n_cycles}"
            )

    # Відправляємо (перевірка 55 хв вже пройдена на початку функції)
    log("Telegram: waiting 60s before sending...")
    time.sleep(60)
    send_telegram("\n".join(lines))
    # Оновлюємо маркер — просто створюємо/перезаписуємо файл
    try:
        with open(telegram_marker, "w") as f:
            f.write(current_time.isoformat())
        log(f"Telegram marker updated: {telegram_marker}")
    except Exception as e:
        log(f"Failed to write telegram marker: {e}")
    
    # Зберігаємо оновлений список
    with open(sent_alerts_file, "w", encoding="utf-8") as f:
        json.dump({
            "last_reset": last_reset_str,
            "alerts": sent_alerts
        }, f, indent=2)

# ── HTML generation ───────────────────────────────────────────────────────────
def fmt_time(dt):   return dt.strftime("%H:%M") if dt else "—"
def eff_color(pct): return "#22c55e" if pct >= 75 else ("#f59e0b" if pct >= 50 else "#ef4444")

def generate_html(cycles, downtimes, period_from, period_to, timeline_data, conn, excel_targets, counter_markers=None):
    generated  = datetime.now().strftime("%d.%m.%Y %H:%M")
    period_str = f"{fmt_time(period_from)} – {fmt_time(period_to)}"
    today_str  = datetime.now().strftime("%Y-%m-%d")
    _gen_hm    = period_to.strftime("%H:%M") if period_to else datetime.now().strftime("%H:%M")

    # ── Stats data ──────────────────────────────────────────────────────
    # Повний список: ALL_MACHINES з конфігу + всі що є в БД + поточні дані
    _all_known_machines = set(ALL_MACHINES) | set(cycles.keys()) | set(downtimes.keys())
    try:
        if conn:
            for _r in conn.execute("SELECT DISTINCT machine FROM daily_summary").fetchall():
                _all_known_machines.add(_canonical_machine(_r[0]))
    except: pass
    _n_known = len(_all_known_machines) if _all_known_machines else 1

    # Годинні дані з БД за останні 7 днів (hourly_stats).
    # save_to_db() перезаписує today, тому current-run теж включено.
    _hourly_cutoff = (datetime.now() - timedelta(days=6)).strftime("%Y-%m-%d")
    _hr_by_date = {}   # {date: {machine: {hour: {"r": run, "t": total}}}}
    try:
        if conn:
            for _r in conn.execute(
                "SELECT date,machine,hour,run_min,total_min FROM hourly_stats "
                "WHERE date >= ? ORDER BY date,machine,hour",
                (_hourly_cutoff,)
            ).fetchall():
                _d, _mraw, _h, _run, _tot = _r
                _m = _canonical_machine(_mraw)
                _hr_by_date.setdefault(_d, {}).setdefault(_m, {})[int(_h)] = {"r": _run or 0, "t": _tot or 0}
    except Exception as _e:
        log(f"hourly_stats load error: {_e}")

    _hdata_all = {}
    for _d2, _per_m in _hr_by_date.items():
        _all_h = set()
        for _m in _per_m: _all_h |= set(_per_m[_m].keys())
        _day_hd = {}
        for _mn in _all_known_machines:
            _day_hd[_mn] = {}
            _mdata = _per_m.get(_mn, {})
            for _h in _all_h:
                _hd = _mdata.get(_h, {"r": 0, "t": 0})
                _day_hd[_mn][str(_h)] = round(_hd["r"]/_hd["t"]*100) if _hd["t"] else 0
        _day_hd["SITE"] = {}
        for _h in _all_h:
            _active_tot = sum(_per_m[_m].get(_h,{}).get("t",0) for _m in _per_m if _per_m[_m].get(_h,{}).get("t",0)>0)
            _n_active = sum(1 for _m in _per_m if _per_m[_m].get(_h,{}).get("t",0)>0)
            _run2 = sum(_per_m[_m].get(_h,{}).get("r",0) for _m in _per_m)
            _avg_t = (_active_tot / _n_active) if _n_active else 0
            _site_denom = _avg_t * _n_known
            _day_hd["SITE"][str(_h)] = min(100, round(_run2/_site_denom*100)) if _site_denom else 0
        _hdata_all[_d2] = _day_hd
    _hourly_js = json.dumps(_hdata_all)
    # Ефективність за робочі години (filtered) для плашки під графіком
    _today_eff = {}
    for _mn in _all_known_machines:
        _dd = downtimes.get(_mn, {})
        _r = _dd.get("total_run", 0)
        _t = _dd.get("total_min", 0)
        _today_eff[_mn] = round(_r / _t * 100) if _t else 0
    _today_eff_js = json.dumps(_today_eff)
    # Денні дані з DB
    _all_daily = {}; _mk_set = set(_all_known_machines)
    try:
        if conn:
            for _r in conn.execute("SELECT date,machine,run_min,down_min,total_min,efficiency FROM daily_summary ORDER BY date").fetchall():
                _d2, _m2_raw, _ru, _dm, _tm, _ef = _r
                _m2 = _canonical_machine(_m2_raw)
                if _d2 not in _all_daily: _all_daily[_d2] = {}
                if _tm or _ru:
                    if _ef and _ef > 0:
                        _all_daily[_d2][_m2] = round(_ef)
                    elif _ru and _ru > 0:
                        # Вихідні або зламані дані — перерахувати з run/(run+down)
                        _denom = (_ru or 0) + (_dm or 0)
                        _all_daily[_d2][_m2] = round(_ru / _denom * 100) if _denom > 0 else 0
                    else:
                        _all_daily[_d2][_m2] = 0
                    _all_daily[_d2].setdefault("__s__",{"r":0})
                    _all_daily[_d2]["__s__"]["r"] += _ru or 0
                _mk_set.add(_m2)
    except: pass
    _n_machines = len(_mk_set)
    for _d2 in _all_daily:
        for _m2 in _mk_set:
            if _m2 not in _all_daily[_d2]:
                _all_daily[_d2][_m2] = 0
        _sd = _all_daily[_d2].pop("__s__",{})
        _ww = _work_window_min(_d2)
        if _ww > 0:
            _site_t = _ww * _n_machines
            _all_daily[_d2]["SITE"] = min(100, round(_sd.get("r",0) / _site_t * 100))
        else:
            # сб/нд: тільки активні станки (робота у вихідні — бонус)
            _active_eff = [v for k,v in _all_daily[_d2].items() if k != "SITE" and v is not None and v > 0]
            _all_daily[_d2]["SITE"] = round(sum(_active_eff) / len(_active_eff)) if _active_eff else 0
    _mk_list  = sorted(_mk_set) + ["SITE"]
    _sk_list  = [(_m.split("_")[0] if "_" in _m else _m) for _m in _mk_list[:-1]] + ["Avg"]
    _col_list = ["#3b82f6","#22c55e","#f59e0b","#ef4444","#a855f7","#06b6d4","#f97316","#ec4899"][:len(_mk_list)]
    _daily_js = json.dumps(_all_daily)
    _mk_js    = json.dumps([str(_m) for _m in _mk_list])
    _sk_js    = json.dumps(_sk_list)
    _col_js   = json.dumps(_col_list)
    # Cycle events for Batch Gantt — останні 30 днів
    _cdata = []
    try:
        if conn:
            _gantt_cutoff = (datetime.now() - timedelta(days=365)).strftime("%Y-%m-%d")
            for _r in conn.execute(
                "SELECT date,machine,program,start_time,end_time,duration "
                "FROM cycle_events WHERE program NOT LIKE 'COUNTER%' AND date >= ? "
                "ORDER BY date,machine,start_time",
                (_gantt_cutoff,)
            ).fetchall():
                _s = _r[3] if _r[3] and _r[3] != "—" else None
                if not _s:
                    continue  # skip records without valid start time (renders at 00:00)
                _e = _r[4] if _r[4] and _r[4] != "—" else None
                if not _e:
                    # Ongoing cycle: use generation time for today, 23:59 for past days
                    if _r[0] == today_str:
                        _e = datetime.now().strftime("%H:%M")
                    else:
                        _e = "23:59"
                _dc = _r[0]  # "2026-04-19" → "260419"
                _d_short = _dc[2:4] + _dc[5:7] + _dc[8:10]
                _s_short = _s.replace(":", "")  # "08:30" → "0830"
                _e_short = _e.replace(":", "") if _e else _e
                # Стискаємо ім'я програми: "WF330-903B.MIN" → "330-903B"
                _prog = _r[2] or ""
                _pu = _prog.upper()
                if _pu.startswith("WF"):
                    _prog = _prog[2:]
                if _pu.endswith(".MIN"):
                    _prog = _prog[:-4]
                _cdata.append({
                    "d": _d_short,
                    "m": (_r[1].split("_")[0] if "_" in _r[1] else _r[1]),
                    "p": _prog, "s": _s_short, "e": _e_short, "dur": _r[5]
                })
    except Exception as _e:
        log(f"  gantt SQL error: {_e}")
    log(f"  gantt: {len(_cdata)} records (last 365 days)")
    _cdata_js = json.dumps(_cdata)
    # ────────────────────────────────────────────────────────────────────

    def timeline_bar(mname):
        segs         = timeline_data.get(mname, [])
        total_sec_tl = max((period_to - period_from).total_seconds(), 1)
        total_min    = total_sec_tl / 60
        VW, VH       = 10000, 44

        def to_x(pct):
            return round(pct / 100 * VW, 2)

        # ── сегменти ──────────────────────────────────────────────────
        rects = ""
        for s in segs:
            color  = "#22c55e" if s["state"] == "1" else "#ef4444"
            seg_id = s["id"]
            label  = s["label"].replace('"', "&quot;").replace("'", "&#39;")
            tip    = f'{s["start"]}\u2013{s["end"]} | {label}'
            x  = to_x(s["x"])
            w  = max(to_x(s["w"]), 0.5)
            rects += (
                f'<rect class="tl-seg" data-id="{seg_id}" data-tip="{tip}" '
                f'x="{x}" y="0" width="{w}" height="{VH}" fill="{color}" cursor="pointer"/>')



        markers_svg = ""
        if counter_markers:
            for ct in counter_markers.get(mname, []):
                ct_sec = (ct - period_from).total_seconds()
                pct = ct_sec / total_sec_tl * 100
                if 0 <= pct <= 100:
                    cx = to_x(pct)
                    markers_svg += (
                        f'<line x1="{cx}" y1="0" x2="{cx}" y2="{VH}" '
                        f'stroke="#a855f7" stroke-width="1" opacity="1" pointer-events="none"/>')

        # ── тіки — JSON масив для JS ───────────────────────────────────
        ticks_json = []
        for i in range(0, int(total_min) + 1, 15):
            pct  = round(i / total_min * 100, 4)
            t    = (period_from + timedelta(minutes=i)).strftime("%H:%M")
            is30 = (i % 30 == 0)
            ticks_json.append({"p": pct, "t": t if is30 else "", "major": is30})

        short  = mname.split("_")[0] if "_" in mname else mname
        uid    = mname.replace(" ", "_").replace("-", "_")

        # SVG — тільки бари, початкова ширина 100%
        svg = (
            f'<svg class="tl-svg" id="svg_{uid}" data-machine="{short}" '
            f'viewBox="0 0 {VW} {VH}" preserveAspectRatio="none" '
            f'style="width:100%;height:{VH}px;display:block;background:#f1f5f9">'
            f'{rects}{markers_svg}'
            f'</svg>')

        # Canvas — тіки, завжди 100% ширини зовнішнього контейнера (не scroll-wrapper)
        # Canvas всередині scroll-wrapper, ширина = SVG ширина при zoom
        canvas = f'<canvas id="tc_{uid}" style="display:block;height:26px"></canvas>'

        ticks_js = (
            f'<script>(function(){{'
            f'var ticks={json.dumps(ticks_json)};'
            f'var cv=document.getElementById("tc_{uid}");'
            f'var svg=document.getElementById("svg_{uid}");'
            f'function drawTicks(w){{'
            f'  var dpr=window.devicePixelRatio||1;'
            f'  cv.style.width=w+"px";'
            f'  cv.width=Math.round(w*dpr);cv.height=Math.round(26*dpr);'
            f'  var ctx=cv.getContext("2d");ctx.scale(dpr,dpr);'
            f'  ctx.clearRect(0,0,w,26);'
            f'  ticks.forEach(function(tk){{'
            f'    var x=tk.p/100*w;'
            f'    ctx.beginPath();ctx.moveTo(x,0);ctx.lineTo(x,tk.major?7:4);'
            f'    ctx.strokeStyle="#cbd5e1";ctx.lineWidth=tk.major?1.5:1;ctx.stroke();'
            f'    if(tk.t){{'
            f'      ctx.font="10px sans-serif";ctx.fillStyle="#64748b";'
            f'      ctx.textAlign="center";ctx.fillText(tk.t,x,18);'
            f'    }}'
            f'  }});'
            f'}}'
            f'var outer=cv.closest(".tl-outer-wrap");'
            f'if(outer) outer._redrawTicks=drawTicks;'
            f'drawTicks(svg.offsetWidth||outer.offsetWidth||800);'
            f'}})();</script>')

        return (
            f'<div class="tl-outer-wrap" style="width:100%">'
            f'<div class="tl-scroll-wrapper" data-machine="{short}" '
            f'style="overflow-x:auto;overflow-y:hidden;width:100%">'
            f'{svg}'
            f'{canvas}'
            f'</div>'
            f'{ticks_js}'
            f'</div>')

    def activity_section(c_list, d_list, mname):
        """Об'єднана таблиця циклів та простоїв, відсортована за часом"""
        segs = timeline_data.get(mname, [])
        
        # Функція пошуку ID для циклів
        def find_cycle_ids(cycle):
            c_start = cycle.get("start")
            if not c_start:
                return ""
            c_end = cycle.get("end") or datetime.now()
            ids = []
            for s in segs:
                if s["state"] != "1":
                    continue
                s_start = datetime.strptime(f"{c_start.strftime('%Y.%m.%d')} {s['start']}", "%Y.%m.%d %H:%M")
                s_end   = datetime.strptime(f"{c_start.strftime('%Y.%m.%d')} {s['end']}", "%Y.%m.%d %H:%M")
                # Округлюємо до хвилини (сегменти таймлайну %H:%M, підцикли до секунди)
                c_end_m   = c_end.replace(second=0, microsecond=0)
                c_start_m = c_start.replace(second=0, microsecond=0)
                if s_start < c_end_m and s_end > c_start_m:
                    ids.append(s["id"])
            return " ".join(ids) if ids else ""
        
        # Функція пошуку ID для простоїв
        def find_down_id(down):
            for s in segs:
                if s["state"] == "0" and s["start"] == down["start"].strftime("%H:%M"):
                    return s["id"]
                if s["state"] == "0" and s["start"] <= down["start"].strftime("%H:%M") <= s["end"]:
                    return s["id"]
            return ""
        
        # Об'єднуємо всі події
        events = []

        for c in c_list:
            if c.get("program", "").upper().startswith("COUNTER"):
                continue
            cycle_start = c["start"] if c.get("start") else datetime.now()
            events.append({
                "type": "cycle",
                "start": cycle_start,
                "end": c.get("end"),
                "duration": c["duration"],
                "cycle_time": c.get("cycle_time"),
                "program": c.get("program", "—"),
                "ongoing": c.get("ongoing", False),
                "ids": find_cycle_ids(c)
            })

        for d in d_list:
            events.append({
                "type": "downtime",
                "start": d["start"],
                "end": d.get("end"),
                "duration": d["duration"],
                "reason": d["reason"],
                "ongoing": d.get("ongoing", False),
                "ids": find_down_id(d)
            })

        events.sort(key=lambda e: e["start"])

        if not events:
            return '<p class="empty">No activity detected</p>'

        # Групуємо: кожен зелений (цикл) + наступні червоні до наступного зеленого = один блок.
        # Один блок = один цикл деталі: виконання програми + заміна деталі.
        # Cycle = cycle_time зеленого рядка (start-to-start, вже розраховано).
        blocks = []
        i = 0
        while i < len(events):
            e = events[i]
            if e["type"] == "cycle":
                block_events = [e]
                j = i + 1
                while j < len(events) and events[j]["type"] != "cycle":
                    block_events.append(events[j])
                    j += 1
                total_block_dur = round(sum(ev.get("duration", 0) for ev in block_events), 2)
                blocks.append({"events": block_events, "cycle_time": total_block_dur})
                i = j
            else:
                # Червоні події до першого зеленого — окремий блок без cycle
                block_events = []
                while i < len(events) and events[i]["type"] != "cycle":
                    block_events.append(events[i])
                    i += 1
                if block_events:
                    blocks.append({"events": block_events, "cycle_time": None})

        rows_html = ""
        for bi, blk in enumerate(blocks):
            blk_events = blk["events"]
            cycle_time = blk["cycle_time"]
            n = len(blk_events)

            if bi > 0:
                rows_html += (
                    f'<tr style="height:3px;padding:0;line-height:0;">'
                    f'<td colspan="6" style="height:3px;padding:0;background:#a855f7;border:none;"></td></tr>'
                )

            for idx, e in enumerate(blk_events):
                badge = ' <span class="badge ongoing">ongoing</span>' if e.get("ongoing") else ""

                if e["type"] == "cycle":
                    icon = "🟢"
                    _prog_disp = e["program"] or "—"
                    if _prog_disp.upper().endswith(".MIN"):
                        _prog_disp = _prog_disp[:-4]
                    detail = _prog_disp
                    row_class = "activity-run"
                else:
                    icon = "🔴"
                    detail = e["reason"]
                    row_class = "activity-down"

                start_s = fmt_time(e["start"])
                end_s = "…" if not e.get("end") else fmt_time(e["end"])

                if idx == 0:
                    if cycle_time is not None:
                        cycle_td = (
                            f'<td rowspan="{n}" style="'
                            f'color:#7c3aed;font-weight:700;'
                            f'text-align:center;vertical-align:middle;'
                            f'white-space:nowrap;border-left:1px solid #000;">'
                            f'{cycle_time} min</td>'
                        )
                    else:
                        cycle_td = f'<td rowspan="{n}" style="border-left:1px solid #000;"></td>'
                else:
                    cycle_td = ""

                rows_html += (
                    f'<tr class="tl-row {row_class}" data-id="{e["ids"]}">'
                    f'<td>{detail}</td>'
                    f'<td>{icon}</td>'
                    f'<td>{start_s}</td>'
                    f'<td>{end_s}{badge}</td>'
                    f'<td><strong>{e["duration"]} min</strong></td>'
                    f'{cycle_td}</tr>'
                )

        # Унікальний ID для цього блоку
        scroll_id = f"activity-scroll-{mname.replace('_', '-')}"
        return (
            f'<div class="resizable-section" id="rs-{scroll_id}" style="height:350px">'
            f'<div class="table-scroll-x" style="height:100%;overflow:hidden"><div class="scroll-table-wrap" style="height:100%">'
            f'<table class="scroll-table"><thead><tr><th>Details</th><th></th><th>Start</th><th>End</th><th>Duration</th><th style="color:#7c3aed;border-left:1px solid #000;text-align:center;">Cycle</th></tr></thead></table>'
            f'<div id="{scroll_id}" class="scroll-tbody-wrap" style="height:calc(100% - 40px);overflow-y:auto">'
            f'<table class="scroll-table"><tbody>{rows_html}</tbody></table>'
            f'</div></div></div></div>'
            f'<script>(function(){{'
            f'var wrap=document.getElementById("{scroll_id}");'
            f'var outer=document.getElementById("rs-{scroll_id}");'
            f'if(!wrap||!outer) return;'
            f'var inner=wrap.querySelector("table");'
            f'if(!inner) return;'
            f'var contentH=inner.offsetHeight+44;'  # 44px = thead
            f'if(contentH<350) outer.style.height=contentH+"px";'
            f'}})();</script>'
        )

    def cycles_section(c_list, mname, excel_targets):
        """Генерує Target Cycle Time з порівнянням з Excel"""
        if not c_list:
            return ""

        # DEBUG: Логуємо скільки targets завантажено
        log(f"cycles_section: machine={mname}, excel_targets count={len(excel_targets)}")
        
        # Витягуємо коротку назву станку (M1, M2 тощо)
        machine_short = mname.split("_")[0] if "_" in mname else mname
        machine_norm  = normalize_program_name(machine_short)

        # Нормалізуємо ключі excel_targets один раз
        excel_norm = [
            (normalize_program_name(p), op, normalize_program_name(m), t, m)
            for (p, op, m), t in excel_targets.items()
        ]

        # Групуємо цикли по програмах (COUNTER.MIN не показуємо)
        by_prog = defaultdict(list)
        for c in c_list:
            prog_name = c["program"] or "—"
            if prog_name.upper().startswith("COUNTER"):
                continue
            by_prog[prog_name].append(c)

        # Для кожної програми використовуємо нову логіку вибірки
        target_rows = []
        for prog, current_cycles in by_prog.items():
            # Визначаємо операцію
            op_num = get_operation_number(prog)

            # Calculated = повний час блоку (run + setup) = start-to-start між
            # послідовними циклами тієї ж програми, без обмеження порогом.
            # Це збігається з колонкою Cycle у Activity Log.
            sorted_cc = sorted(current_cycles, key=lambda c: c["start"])
            block_times = []
            for idx_c, cc in enumerate(sorted_cc):
                if idx_c + 1 < len(sorted_cc) and cc.get("start") and sorted_cc[idx_c + 1].get("start"):
                    bt = round((sorted_cc[idx_c + 1]["start"] - cc["start"]).total_seconds() / 60, 2)
                    if bt > 0:
                        block_times.append(bt)
                else:
                    # Останній цикл — використовуємо cycle_time або duration
                    ct = cc.get("cycle_time") or cc.get("duration", 0)
                    if ct and ct > 0:
                        block_times.append(ct)

            if block_times:
                calc_target = calculate_real_cycle_time(block_times)
            else:
                calc_target = None

            if calc_target is None:
                continue

            info_text = f"{len(current_cycles)} cycles today"

            # Шукаємо Excel Target з урахуванням станку та операції
            excel_target = None
            prog_normalized = normalize_program_name(prog)
            found_for_other_machine = None

            for ep_norm, eop, em_norm, time_val, em_orig in excel_norm:
                if prog_normalized == ep_norm and op_num == eop:
                    if machine_norm == em_norm:
                        excel_target = time_val
                        break
                    else:
                        found_for_other_machine = em_orig
            
            # Порівняння
            if excel_target:
                diff = round(calc_target - excel_target, 2)  # Округлення до сотих
                diff_pct = round((diff / excel_target) * 100, 2) if excel_target else 0
                sign = f'+{diff}' if diff > 0 else str(diff)
                sign_pct = f'+{diff_pct}' if diff_pct > 0 else str(diff_pct)
                if abs(diff_pct) <= 5:
                    comparison = f'<span style="color:#22c55e">{sign} min ({sign_pct}%)</span>'
                elif diff > 0:
                    comparison = f'<span style="color:#ef4444">{sign} min ({sign_pct}%)</span>'
                else:
                    comparison = f'<span style="color:#3b82f6">{sign} min ({sign_pct}%)</span>'
                excel_col = f'{excel_target} min'
            else:
                # Якщо не знайшли для цього станку, але є для іншого
                if found_for_other_machine:
                    comparison = '<span style="color:#f59e0b">Wrong machine</span>'
                    excel_col = f'No norm for {machine_short}'
                else:
                    comparison = '<span style="color:#94a3b8">No data</span>'
                    excel_col = '—'
            
            target_rows.append(
                f'<tr><td><em>{prog[:-4] if prog.upper().endswith(".MIN") else prog}</em></td>'
                f'<td><span class="badge" style="background:#3F51B5;color:white;padding:2px 8px;border-radius:2px">OP{op_num}</span></td>'
                f'<td>{len(current_cycles)}</td>'
                f'<td><strong>{calc_target} min</strong></td>'
                f'<td>{excel_col}</td>'
                f'<td>{comparison}</td>'
                f'<td style="font-size:0.75rem;color:#64748b">{info_text}</td></tr>'
            )

        return (
            f'<div class="section-title" style="border-top:2px dashed #e2e8f0">🎯 Target Cycle Time</div>'
            f'<div class="table-scroll-x"><table><thead><tr><th>Program</th><th>OP</th><th>Total</th><th>Calculated</th><th>Target</th><th>Difference</th><th>Info</th></tr></thead>'
            f'<tbody>{"".join(target_rows)}</tbody></table></div>'
        )

    machines_html = ""
    machine_names = sorted(cycles.keys())
    nav_buttons = (
        '<div style="border-top:1px solid #475569;margin:4px 0"></div>\n'
    ) + "".join(
        f'<a href="#machine-{mn.split("_")[0] if "_" in mn else mn}" class="nav-btn">{mn.split("_")[0] if "_" in mn else mn}</a>\n'
        for mn in machine_names
    )
    for i, mname in enumerate(machine_names):
        c_list     = cycles.get(mname, [])
        d_data     = downtimes.get(mname, {})
        d_list     = d_data.get("downtimes", [])
        total_min  = d_data.get("total_min", 1)
        total_run  = d_data.get("total_run", 0)
        total_down = d_data.get("total_down", 0)
        eff        = round(total_run / total_min * 100) if total_min else 0
        short_name = mname.split("_")[0] if "_" in mname else mname

        sep = f'<div class="machine-sep">{short_name}</div>' if i > 0 else ''
        machines_html += sep + f"""
        <div class="machine-card" id="machine-{short_name}">
          <div class="machine-header">
            <div class="machine-title">
              <span class="machine-id">{short_name}</span>
              <span class="machine-full">{mname}</span>
            </div>
            <div class="eff-badge" style="background:{eff_color(eff)}">
              Efficiency: {eff}%
              <span class="eff-detail">({total_run} / {total_min} min)</span>
            </div>
          </div>
          <div class="section-title">⏱ Timeline</div>
          <div style="padding:10px 20px 4px">{timeline_bar(mname)}</div>
          <div class="section-title">📋 Activity Log — {len(c_list)} cycles, {len(d_list)} downtimes ({total_down} min)</div>
          <div style="padding:0 0 4px">{activity_section(c_list, d_list, mname)}</div>
          {cycles_section(c_list, mname, excel_targets)}
        </div>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Machine Report — {generated}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Roboto','Segoe UI',Arial,sans-serif;background:#ffffff;color:#212121;font-size:15px;margin:0;-webkit-font-smoothing:antialiased;-moz-osx-font-smoothing:grayscale;text-rendering:optimizeLegibility}}

  /* ── Header ── */
  .header{{background:#1450CF;color:white;padding:10px 20px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;box-shadow:0 2px 4px rgba(0,0,0,.2)}}
  .header h1{{font-size:1.2rem;font-weight:500}}
  .header .meta{{font-size:.8rem;opacity:.9;text-align:right}}
  .header-tabs{{display:flex;gap:4px;align-items:center}}

  /* ── Layout ── */
  .container{{max-width:1100px;margin:16px auto;padding:0 12px}}

  /* ── Machine card ── */
  .machine-card{{background:#FFFFFF;border-radius:2px;box-shadow:0 2px 4px rgba(0,0,0,.3);margin-bottom:16px;overflow:hidden}}
  .machine-header{{background:#1e293b;color:white;padding:14px 16px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px}}
  .machine-title{{display:flex;flex-direction:column;gap:2px}}
  .machine-id{{font-size:1.1rem;font-weight:500}}
  .machine-full{{font-size:.7rem;opacity:.8;word-break:break-all}}
  .eff-badge{{padding:5px 12px;border-radius:2px;font-weight:500;font-size:.85rem;color:white;white-space:nowrap}}
  .eff-detail{{font-size:.72rem;font-weight:400;opacity:.9;margin-left:4px}}

  /* ── Section title ── */
  .section-title{{padding:10px 16px 5px;font-weight:500;font-size:.8rem;color:#757575;border-top:1px solid #E0E0E0;text-transform:uppercase;letter-spacing:.05em}}

  /* ── Tables (desktop) ── */
  table{{width:100%;border-collapse:collapse;font-size:.85rem}}
  th{{background:#F8F9FA;padding:8px 12px;text-align:left;font-weight:500;color:#495057;border-bottom:2px solid #DEE2E6;white-space:nowrap}}
  td{{padding:8px 12px;border-bottom:1px solid #DEE2E6;word-break:break-word}}
  tr:last-child td{{border-bottom:none}}
  tr:hover td{{background:#E9ECEF}}
  .tl-row{{cursor:pointer;transition:background-color 0.2s ease}}
  .tl-row:hover td{{background:#E3F2FD!important}}
  .activity-run td{{background:#E8F5E9}}
  .activity-down td{{background:#FFEBEE}}
  .activity-run:hover td{{background:#C8E6C9!important}}
  .activity-down:hover td{{background:#FFCDD2!important}}

  /* ── Scrollable tables ── */
  .scroll-table-wrap{{width:100%;border-bottom:1px solid #e2e8f0}}
  .scroll-table{{width:100%;border-collapse:collapse;table-layout:fixed;font-size:.85rem}}
  .scroll-table th,.scroll-table td{{padding:8px 12px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
  .scroll-table thead th{{background:#F8F9FA;font-weight:500;color:#495057;border-bottom:2px solid #DEE2E6;position:sticky;top:0;z-index:1}}
  .scroll-table tbody tr:last-child td{{border-bottom:none}}
  .scroll-table tbody tr:hover td{{background:#E9ECEF}}
  .scroll-tbody-wrap{{display:block}}
  .scroll-tbody-wrap::-webkit-scrollbar{{width:5px}}
  .scroll-tbody-wrap::-webkit-scrollbar-track{{background:#F8F9FA}}
  .scroll-tbody-wrap::-webkit-scrollbar-thumb{{background:#ADB5BD;border-radius:2px}}
  
  /* ── Resizable Activity Log ── */
  .resizable-section{{
    position:relative;
    overflow:auto;
    border:1px solid #e2e8f0;
    border-radius:6px;
    background:#fff;
    resize:vertical;
    min-height:150px;
  }}
  .resizable-section::-webkit-resizer{{
    background:#cbd5e1;
    border-radius:3px;
  }}

  /* ── Timeline ── */
  .tl-scroll-wrapper{{overflow-x:auto;overflow-y:hidden;position:relative;margin:0 -12px;padding:0 12px}}
  .tl-scroll-wrapper::-webkit-scrollbar{{height:8px}}
  .tl-scroll-wrapper::-webkit-scrollbar-track{{background:#f1f5f9}}
  .tl-scroll-wrapper::-webkit-scrollbar-thumb{{background:#cbd5e1;border-radius:4px}}
  .tl-svg{{display:block;border-radius:4px;user-select:none;cursor:grab}}

  /* ── Legend ── */
  .legend{{display:flex;gap:14px;padding:4px 16px 10px;font-size:.78rem;flex-wrap:wrap;color:#212121}}
  .legend span{{display:flex;align-items:center;gap:4px}}
  .dot{{width:11px;height:11px;border-radius:2px;display:inline-block;flex-shrink:0}}

  /* ── Misc ── */
  .badge{{display:inline-block;padding:2px 7px;border-radius:10px;font-size:.72rem;font-weight:600;margin-left:3px}}
  .badge.ongoing{{background:#dbeafe;color:#1d4ed8}}
  .empty{{padding:12px 16px;color:#94a3b8;font-style:italic;font-size:.85rem}}
  .footer{{text-align:center;color:#94a3b8;font-size:.75rem;padding:16px;word-break:break-all}}

  /* ── Tooltip ── */
  #tl-tooltip{{
    position:fixed;pointer-events:none;z-index:9999;
    background:#1e293b;color:white;
    padding:6px 10px;border-radius:6px;font-size:.78rem;
    box-shadow:0 4px 12px rgba(0,0,0,.3);
    display:none;max-width:240px;white-space:normal;line-height:1.4
  }}

  /* ── Highlight ── */
  .tl-seg.dim{{opacity:.25}}
  .tl-seg.highlight{{opacity:1;filter:brightness(1.18)}}
  .tl-row{{transition:background-color 0.2s ease}}
  .tl-row.highlight td{{background:#fde047!important;font-weight:600}}
  .tl-row:hover td{{background:#f0f9ff!important}}

  /* ══════════════════════════════════════════
     MOBILE  ≤ 600px
  ══════════════════════════════════════════ */
  @media(max-width:600px){{
    body{{font-size:14px}}
    .header{{padding:10px 12px}}
    .header h1{{font-size:0.95rem}}
    .header-tabs{{gap:3px}}
    .tab-btn{{padding:4px 10px;font-size:0.75rem}}
    .container{{padding:0 6px;margin:8px auto}}
    .machine-card{{border-radius:6px;margin-bottom:10px}}
    /* machine-header — стак вертикально */
    .machine-header{{padding:10px 12px;flex-direction:column;align-items:flex-start;gap:6px}}
    .eff-badge{{font-size:.8rem;padding:4px 10px;align-self:stretch;text-align:center}}
    .eff-detail{{display:none}}
    .section-title{{padding:7px 12px 4px;font-size:.72rem}}
    /* Таблиці */
    .table-scroll-x{{overflow-x:auto;-webkit-overflow-scrolling:touch}}
    table,.scroll-table{{font-size:.75rem;min-width:300px}}
    th,td,.scroll-table th,.scroll-table td{{padding:6px 8px}}
    /* Resizable — прибираємо resize handle на touch */
    .resizable-section{{resize:none}}
    /* Tooltip */
    #tl-tooltip{{
      position:fixed;bottom:16px;left:50%;transform:translateX(-50%);
      top:auto!important;max-width:92vw;text-align:center
    }}
    .legend{{padding:4px 10px 8px;font-size:.72rem;gap:8px}}
    .footer{{font-size:.7rem;padding:10px 10px 20px}}
    /* Stats */
    .chart-wrap{{height:220px}}
    .chart-panel{{padding:12px 10px}}
    .stats-controls{{gap:6px}}
    .stats-controls label{{font-size:0.78rem}}
    .stats-controls input[type=date]{{font-size:0.78rem;padding:4px 6px;max-width:130px}}
    .stats-controls button{{padding:5px 10px;font-size:0.75rem}}
    .stats-table th,.stats-table td{{padding:4px 4px;font-size:0.75rem;white-space:nowrap;overflow:hidden}}
    .chart-legend{{gap:6px}}
    .leg-item{{font-size:0.72rem}}
    /* Scroll-top */
    #scroll-top{{right:12px;left:auto;bottom:20px;width:40px;height:40px;font-size:1.2rem}}
  }}
  /* Nav drawer (all screens) */
  .nav-sidebar{{
    position:fixed;top:0;left:0;bottom:0;width:fit-content;
    transform:translateX(-100%);
    transition:transform .25s ease;
    display:flex;flex-direction:column;
    background:transparent;padding:60px 8px 16px;gap:6px;
    z-index:1500;overflow-y:auto;
  }}
  .nav-sidebar.open{{transform:translateX(0)}}
  #nav-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:1400;}}
  #nav-overlay.open{{display:block}}
  #nav-toggle{{
    display:flex;position:fixed;left:0;top:50%;transform:translateY(-50%);
    width:26px;height:52px;background:#1e293b;color:#fff;
    border:none;border-radius:0 8px 8px 0;font-size:1rem;cursor:pointer;
    z-index:1600;align-items:center;justify-content:center;
    box-shadow:2px 0 8px rgba(0,0,0,.35);
  }}
  .nav-btn{{display:block;padding:14px 10px;font-size:1.1rem;white-space:nowrap;border-radius:10px;width:100%;text-align:center;background:#1e293b;color:#fff;font-weight:800;text-decoration:none;box-shadow:0 2px 8px rgba(0,0,0,0.35);transition:background .15s,box-shadow .15s;letter-spacing:0.04em}}
  .nav-btn:hover{{background:#3b82f6;box-shadow:0 4px 14px rgba(59,130,246,0.4)}}
  .nav-tab-btn{{background:#1450CF!important;color:#fff!important}}
  .nav-tab-btn.active{{background:#3b82f6!important;box-shadow:0 0 0 2px #fff}}
  /* scroll top */
  #scroll-top{{position:fixed;bottom:24px;left:8px;width:44px;height:44px;border-radius:50%;background:#1e293b;color:#fff;border:none;font-size:1.4rem;cursor:pointer;box-shadow:0 4px 12px rgba(0,0,0,.3);display:none;align-items:center;justify-content:center;z-index:2000;transition:background .15s}}
  #scroll-top:hover{{background:#3b82f6}}
  #scroll-top.visible{{display:flex}}
  /* stats */
  .stats-controls{{display:flex;align-items:center;gap:12px;flex-wrap:wrap;padding:12px 0}}
  .stats-controls label{{font-size:0.85rem;color:#475569;display:flex;align-items:center;gap:6px}}
  .stats-controls input[type=date]{{padding:5px 8px;border:1px solid #cbd5e1;border-radius:6px;font-size:0.85rem}}
  .stats-controls button{{padding:6px 14px;background:#1e293b;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:0.82rem;font-weight:600}}
  .stats-controls button:hover{{background:#3b82f6}}
  .two-charts{{display:flex;flex-direction:column;gap:20px;margin-top:12px}}
  .chart-panel{{background:#fff;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.1);padding:16px}}
  .chart-title{{font-size:0.9rem;font-weight:700;color:#1e293b;margin-bottom:10px}}
  .chart-scroll-outer{{overflow-x:auto;-webkit-overflow-scrolling:touch;border:1px solid red;margin-left:-16px;margin-right:-16px}}
  .chart-wrap{{position:relative;height:350px;min-width:600px;background:#f8fafc}}
  .chart-legend{{display:flex;flex-wrap:wrap;gap:10px;margin-top:8px}}
  .leg-item{{display:flex;align-items:center;gap:5px;font-size:0.8rem;color:#334155;font-weight:500}}
  .leg-dot{{width:11px;height:11px;border-radius:50%;display:inline-block}}
  .stats-table{{border-collapse:collapse;font-size:0.96rem;width:100%;margin-top:16px;table-layout:fixed}}
  .stats-table th{{background:#1e293b;color:#fff;padding:6px 12px;text-align:center}}
  .stats-table td{{padding:5px 12px;text-align:center;border-bottom:1px solid #e2e8f0}}
  @media(max-width:900px){{.two-charts{{grid-template-columns:1fr}}}}
  /* ── Block dividers ── */
  .block-header{{background:#1e293b;color:#fff;padding:10px 16px;border-radius:6px;margin:20px 0 12px;display:flex;align-items:center;gap:8px;font-size:0.9rem;font-weight:700;letter-spacing:0.04em}}
  .block-header.machines{{background:#1e293b;margin-top:28px}}
  /* ── Machine separator ── */
  .machine-sep{{display:flex;align-items:center;gap:10px;margin:24px 0 10px;font-size:0.72rem;color:#94a3b8;font-weight:700;letter-spacing:0.1em;text-transform:uppercase}}
  .machine-sep::before,.machine-sep::after{{content:'';flex:1;height:1px;background:#e2e8f0}}
  .machine-card{{border-top:3px solid #3DA9D7}}
</style>
</head>
<body>
<div class="header">
  <h1>📊 Machine Report</h1>
  <div class="meta">
    Period: {period_str}<br>
    Generated: {generated}
  </div>
</div>
<div class="container">
  <div class="block-header">📊 Statistics</div>
  <div id="machine-filter" style="display:flex;flex-wrap:wrap;gap:6px;padding:6px 0 8px"></div>
  <div class="two-charts">
    <div class="chart-panel">
      <h3 class="chart-title" id="today-chart-title">Today — Hourly Efficiency</h3>
      <div id="today-day-selector" style="display:flex;gap:6px;flex-wrap:wrap;padding:4px 0 10px"></div>
      <div class="chart-scroll-outer"><div class="chart-wrap"><canvas id="effChartToday"></canvas></div></div>
      <div id="today-table-wrap"></div>
    </div>
    <div class="chart-panel">
      <h3 class="chart-title">Period Trend</h3>
      <div class="stats-controls">
        <label>From: <input type="text" id="stat-from" placeholder="dd.mm.yyyy" value="{(datetime.now()-timedelta(days=6)).strftime('%d.%m.%Y')}" style="width:90px"></label>
        <label>To: <input type="text" id="stat-to" placeholder="dd.mm.yyyy" value="{datetime.now().strftime('%d.%m.%Y')}" style="width:90px"></label>
        <button onclick="updatePeriodChart()">Apply</button>
        <button onclick="setRange(7)">7d</button>
        <button onclick="setRange(30)">30d</button>
        <button onclick="setRange(90)">90d</button>
        <button onclick="setRange(180)">180d</button>
        <button onclick="setRange(365)">1y</button>
      </div>
      <div class="chart-scroll-outer">
        <div class="chart-wrap"><canvas id="effChartPeriod"></canvas></div>
      </div>
      <div style="margin-top:6px;font-size:0.75rem;font-weight:700;color:#475569;padding:4px 0 2px;text-transform:uppercase;letter-spacing:.05em">Batch Gantt</div>
      <div id="gantt-scroll-outer" class="chart-scroll-outer">
        <div class="chart-wrap"><canvas id="batchGantt" style="display:block"></canvas></div>
      </div>
    </div>
  </div>
  <div id="batch-tooltip" style="position:fixed;pointer-events:none;z-index:9999;background:#1e293b;color:white;padding:6px 10px;border-radius:6px;font-size:.78rem;box-shadow:0 4px 12px rgba(0,0,0,.3);display:none;max-width:240px;white-space:normal;line-height:1.4"></div>
  <div class="block-header machines">🏭 Machines
    <span style="margin-left:auto;font-weight:400;font-size:0.78rem;opacity:0.8">
      <span class="dot" style="background:#4CAF50;border-radius:2px"></span> Running &nbsp;
      <span class="dot" style="background:#F44336;border-radius:2px"></span> Downtime
    </span>
  </div>
  {machines_html}
</div>
<div id="nav-overlay"></div>
<button id="nav-toggle" title="Навігація">&#9776;</button>
<div class="nav-sidebar" id="nav-sidebar">
{nav_buttons}
</div>
<button id="scroll-top" onclick="window.scrollTo({{top:0,behavior:'smooth'}})" title="↑">↑</button>
<div class="footer">Source: Connect Plan WebAPI ({API_BASE}) &nbsp;|&nbsp; DB: {DB_FILE}</div>
<div id="tl-tooltip"></div>
<script>
function localISO(d){{var y=d.getFullYear(),m=d.getMonth()+1,dd=d.getDate();return y+'-'+(m<10?'0':'')+m+'-'+(dd<10?'0':'')+dd;}}
(function(){{
  var tip = document.getElementById("tl-tooltip");

  // ── Tooltip + highlight при наведенні на сегмент таймлайну ────────────────
  document.querySelectorAll(".tl-seg").forEach(function(seg){{
    seg.addEventListener("mouseenter", function(e){{
      var id=seg.dataset.id, txt=seg.dataset.tip;
      tip.textContent=txt; tip.style.display="block";
      var svg=seg.closest("svg");
      svg.querySelectorAll(".tl-seg").forEach(function(s){{ s.classList.add("dim"); }});
      seg.classList.remove("dim"); seg.classList.add("highlight");
      if(id) document.querySelectorAll('.tl-row').forEach(function(r){{
        var ids=r.dataset.id?r.dataset.id.split(' '):[];
        if(ids.indexOf(id)!==-1) r.classList.add("highlight");
      }});
    }});
    seg.addEventListener("mousemove", function(e){{
      tip.style.left=(e.clientX+14)+"px"; tip.style.top=(e.clientY-32)+"px";
    }});
    seg.addEventListener("mouseleave", function(){{
      tip.style.display="none";
      var svg=seg.closest("svg");
      svg.querySelectorAll(".tl-seg").forEach(function(s){{ s.classList.remove("dim","highlight"); }});
      var id=seg.dataset.id;
      if(id) document.querySelectorAll('.tl-row').forEach(function(r){{
        var ids=r.dataset.id?r.dataset.id.split(' '):[];
        if(ids.indexOf(id)!==-1) r.classList.remove("highlight");
      }});
    }});
    seg.addEventListener("click", function(){{
      var id=seg.dataset.id; if(!id) return;
      var targetRow=null;
      document.querySelectorAll('.tl-row').forEach(function(r){{
        var ids=r.dataset.id?r.dataset.id.split(' '):[];
        if(ids.indexOf(id)!==-1) targetRow=r;
      }});
      if(!targetRow) return;
      var sc=targetRow.closest(".scroll-tbody-wrap");
      if(sc){{ var rr=targetRow.getBoundingClientRect(),cr=sc.getBoundingClientRect(); sc.scrollTo({{top:sc.scrollTop+(rr.top-cr.top)-(cr.height/2)+(rr.height/2),behavior:'smooth'}}); }}
      else {{ window.scrollTo({{top:targetRow.getBoundingClientRect().top+window.scrollY-window.innerHeight/2,behavior:'smooth'}}); }}
      targetRow.classList.add("highlight");
      setTimeout(function(){{ targetRow.classList.remove("highlight"); }},1500);
    }});
    var _tt,_tm=false;
    seg.addEventListener("touchstart",function(e){{
      _tt=Date.now();_tm=false;e.preventDefault();
      tip.textContent=seg.dataset.tip;tip.style.display="block";
      var svg=seg.closest("svg");
      svg.querySelectorAll(".tl-seg").forEach(function(s){{s.classList.add("dim");}});
      seg.classList.remove("dim");seg.classList.add("highlight");
      var id=seg.dataset.id;
      if(id) document.querySelectorAll('.tl-row').forEach(function(r){{
        var ids=r.dataset.id?r.dataset.id.split(' '):[];
        if(ids.indexOf(id)!==-1) r.classList.add("highlight");
      }});
    }},{{passive:false}});
    seg.addEventListener("touchmove",function(){{_tm=true;}});
    seg.addEventListener("touchend",function(){{
      var id=seg.dataset.id;
      if(!_tm&&(Date.now()-_tt)<300&&id){{
        var tr=null;
        document.querySelectorAll('.tl-row').forEach(function(r){{
          var ids=r.dataset.id?r.dataset.id.split(' '):[];
          if(ids.indexOf(id)!==-1) tr=r;
        }});
        if(tr){{
          var sc=tr.closest(".scroll-tbody-wrap");
          if(sc){{var rr=tr.getBoundingClientRect(),cr=sc.getBoundingClientRect();sc.scrollTo({{top:sc.scrollTop+(rr.top-cr.top)-(cr.height/2),behavior:'smooth'}});}}
          else{{window.scrollTo({{top:tr.getBoundingClientRect().top+window.scrollY-window.innerHeight/2,behavior:'smooth'}});}}
          tr.classList.add("highlight");setTimeout(function(){{tr.classList.remove("highlight");}},1500);
        }}
      }}
      setTimeout(function(){{
        tip.style.display="none";
        var svg=seg.closest("svg");
        svg.querySelectorAll(".tl-seg").forEach(function(s){{s.classList.remove("dim","highlight");}});
        if(id) document.querySelectorAll('.tl-row').forEach(function(r){{
          var ids=r.dataset.id?r.dataset.id.split(' '):[];
          if(ids.indexOf(id)!==-1) r.classList.remove("highlight");
        }});
      }},1400);
    }});
  }});

  // ── Highlight при наведенні на рядок таблиці ──────────────────────
  document.querySelectorAll(".tl-row").forEach(function(row){{
    row.addEventListener("mouseenter",function(){{
      var ids=row.dataset.id?row.dataset.id.split(' '):[];
      if(!ids.length) return;
      var svg=null,segs=[];
      ids.forEach(function(id){{
        var seg=document.querySelector('.tl-seg[data-id="'+id+'"]');
        if(seg){{if(!svg)svg=seg.closest("svg");segs.push(seg);}}
      }});
      if(!segs.length) return;
      svg.querySelectorAll(".tl-seg").forEach(function(s){{s.classList.add("dim");}});
      segs.forEach(function(s){{s.classList.remove("dim");s.classList.add("highlight");}});
      tip.textContent=segs[0].dataset.tip;tip.style.display="block";
      tip.style.left=(row.getBoundingClientRect().right+10)+"px";
      tip.style.top=(row.getBoundingClientRect().top+window.scrollY)+"px";
    }});
    row.addEventListener("mouseleave",function(){{
      var ids=row.dataset.id?row.dataset.id.split(' '):[];
      ids.forEach(function(id){{
        var seg=document.querySelector('.tl-seg[data-id="'+id+'"]');
        if(seg){{var svg=seg.closest("svg");svg.querySelectorAll(".tl-seg").forEach(function(s){{s.classList.remove("dim","highlight");}});}}
      }});
      tip.style.display="none";
    }});
    row.addEventListener("click",function(){{
      var ids=row.dataset.id?row.dataset.id.split(' '):[];
      if(!ids.length) return;
      var fs=document.querySelector('.tl-seg[data-id="'+ids[0]+'"]');
      if(!fs) return;
      var wrapper=fs.closest(".tl-scroll-wrapper");
      if(!wrapper) return;
      var sr=fs.getBoundingClientRect(),wr=wrapper.getBoundingClientRect();
      wrapper.scrollTo({{left:wrapper.scrollLeft+(sr.left-wr.left)-(wr.width/2)+(sr.width/2),behavior:'smooth'}});
    }});
  }});

  // ── Drag to scroll ─────────────────────────────────────────────────
  document.querySelectorAll(".tl-scroll-wrapper").forEach(function(wrapper){{
    var svg=wrapper.querySelector("svg");if(!svg) return;
    var isDown=false,startX,slStart,isDragging=false;
    svg.addEventListener("mousedown",function(e){{isDown=true;isDragging=false;startX=e.pageX;slStart=wrapper.scrollLeft;wrapper.style.cursor="grabbing";}});
    document.addEventListener("mousemove",function(e){{if(!isDown) return;e.preventDefault();isDragging=true;wrapper.scrollLeft=slStart+(startX-e.pageX)*2;}});
    document.addEventListener("mouseup",function(){{if(isDown){{isDown=false;wrapper.style.cursor="";setTimeout(function(){{isDragging=false;}},50);}}}});
    svg.addEventListener("click",function(e){{if(isDragging){{e.stopPropagation();e.preventDefault();}}}},true);
    var txStart,tsLeft;
    svg.addEventListener("touchstart",function(e){{txStart=e.touches[0].pageX;tsLeft=wrapper.scrollLeft;}},{{passive:true}});
    svg.addEventListener("touchmove",function(e){{if(!txStart) return;wrapper.scrollLeft=tsLeft+(txStart-e.touches[0].pageX)*2;}},{{passive:true}});
    svg.addEventListener("touchend",function(){{txStart=null;}});
  }});

  // ── Ctrl+Wheel zoom ─────────────────────────────────────────────────
  // Zoom: SVG розтягується (бари), canvas перемальовується (тіки завжди чіткі)
  document.querySelectorAll(".tl-outer-wrap").forEach(function(outer){{
    var wrapper=outer.querySelector(".tl-scroll-wrapper");
    var svg=wrapper?wrapper.querySelector("svg"):null;
    if(!svg) return;
    var BASE=wrapper.offsetWidth||800,cur=BASE,MIN=BASE,MAX=BASE*20;
    wrapper.addEventListener("wheel",function(e){{
      if(!e.ctrlKey) return;e.preventDefault();
      var rect=svg.getBoundingClientRect();
      var nw=Math.round(Math.min(MAX,Math.max(MIN,cur*(e.deltaY<0?1.15:1/1.15))));
      if(nw===cur) return;
      var ratio=(e.clientX-rect.left+wrapper.scrollLeft)/cur;
      cur=nw;svg.style.width=nw+"px";
      wrapper.scrollLeft=Math.round(ratio*nw-(e.clientX-rect.left));
      // Canvas тіки: передаємо точну ширину SVG
      if(outer._redrawTicks) outer._redrawTicks(nw);
    }},{{passive:false}});
  }});
}})();

// ── Stats charts ──────────────────────────────────────────────────
(function(){{
  var ALL   = {_daily_js};
  var HDATA = {_hourly_js};
  var TEFF  = {_today_eff_js};
  var MK    = {_mk_js};
  var SK    = {_sk_js};
  var COLS  = {_col_js};
  var REPORT_HM = "{_gen_hm}";
  var tCh=null, pCh=null;

  // Plugin: тонка червона вертикальна лінія у момент генерації звіту.
  // Малюється на позиції лейблу REPORT_HM (напр. "14:28"), який окремо додано в labels —
  // тому криві дотягуються до неї, а нативний tooltip працює як на звичайних точках.
  var reportLinePlugin = {{
    id: 'reportLine',
    afterDatasetsDraw: function(chart) {{
      if (!REPORT_HM) return;
      var iLine = chart.data.labels.indexOf(REPORT_HM);
      if (iLine < 0) return;
      var xLine = Math.round(chart.scales.x.getPixelForValue(iLine)) + 0.5;
      var ca = chart.chartArea;
      var ctx = chart.ctx;
      ctx.save();
      ctx.strokeStyle = '#ef4444';
      ctx.lineWidth = 1;
      ctx.setLineDash([4,3]);
      ctx.beginPath();
      ctx.moveTo(xLine, ca.top);
      ctx.lineTo(xLine, ca.bottom);
      ctx.stroke();
      ctx.restore();
    }}
  }};

  // ── Machine filter checkboxes ─────────────────────────────────────
  (function(){{
    var fd=document.getElementById('machine-filter');
    if(!fd) return;
    function setAll(on){{
      var cbs=fd.querySelectorAll('input[type=checkbox]');
      cbs.forEach(function(cb){{
        cb.checked=on;
        var lbl=cb.parentElement;
        lbl.style.background=on?(lbl.dataset.bg||'transparent'):'transparent';
      }});
      if(tCh) initToday();
      updatePeriodChart();
    }}
    var btnStyle='cursor:pointer;font-size:0.75rem;font-weight:700;padding:4px 10px;'
      +'border-radius:20px;border:2px solid #cbd5e1;background:#fff;color:#1e293b;'
      +'transition:background .15s';
    var btnAll=document.createElement('button');
    btnAll.type='button';btnAll.textContent='All';btnAll.style.cssText=btnStyle;
    btnAll.addEventListener('click',function(){{setAll(true);}});
    var btnNone=document.createElement('button');
    btnNone.type='button';btnNone.textContent='None';btnNone.style.cssText=btnStyle;
    btnNone.addEventListener('click',function(){{setAll(false);}});
    fd.appendChild(btnAll);fd.appendChild(btnNone);
    MK.forEach(function(k,i){{
      var lbl=document.createElement('label');
      lbl.dataset.bg=COLS[i]+'22';
      lbl.style.cssText='display:flex;align-items:center;gap:4px;cursor:pointer;font-size:0.8rem;'
        +'color:#1e293b;font-weight:600;user-select:none;padding:3px 10px 3px 6px;'
        +'border-radius:20px;border:2px solid '+COLS[i]+';background:'+COLS[i]+'22;transition:background .15s';
      var cb=document.createElement('input');
      cb.type='checkbox';cb.checked=true;cb.dataset.m=k;
      cb.style.cssText='cursor:pointer;accent-color:'+COLS[i]+';width:13px;height:13px;flex-shrink:0';
      cb.addEventListener('change',function(){{
        lbl.style.background=cb.checked?COLS[i]+'22':'transparent';
        if(tCh) initToday();
        updatePeriodChart();
      }});
      lbl.appendChild(cb);
      var dot=document.createElement('span');
      dot.style.cssText='width:9px;height:9px;border-radius:50%;background:'+COLS[i]+';display:inline-block;flex-shrink:0';
      lbl.appendChild(dot);
      lbl.appendChild(document.createTextNode('\u00a0'+SK[i]));
      fd.appendChild(lbl);
    }});
  }})();

  function getSelected(){{
    var cbs=document.querySelectorAll('#machine-filter input[type=checkbox]');
    var sel=[];cbs.forEach(function(cb){{if(cb.checked)sel.push(cb.dataset.m);}});
    return sel;
  }}
  function ds(labels,getFn){{
    var sel=getSelected();
    return MK.filter(function(k){{return sel.indexOf(k)!==-1;}}).map(function(k){{
      var i=MK.indexOf(k);
      var data;
      if(k==='SITE'){{
        var selM=sel.filter(function(m){{return m!=='SITE';}});
        data=labels.map(function(l){{
          if(!selM.length) return null;
          var sum=selM.reduce(function(s,m){{var v=getFn(m,l);return s+(v!=null?v:0);}},0);
          return Math.round(sum/selM.length);
        }});
      }}else{{
        data=labels.map(function(l){{var v=getFn(k,l);return v!=null?v:0;}});
      }}
      return {{label:SK[i],data:data,
        borderColor:COLS[i],backgroundColor:COLS[i]+'22',tension:0.3,
        pointRadius:5,pointHoverRadius:7,borderWidth:k==='SITE'?4:1.5,
        borderDash:k==='SITE'?[8,4]:[],
        fill:false,spanGaps:false}};
    }});
  }}
  function yMax(datasets){{
    var mx=100;
    datasets.forEach(function(d){{d.data.forEach(function(v){{if(v!=null&&v>mx)mx=v;}});}});
    return Math.ceil(mx/10)*10+10;
  }}
  function opts(datasets){{var ym=datasets?yMax(datasets):110;return{{clip:false,responsive:true,maintainAspectRatio:false,layout:{{padding:0}},
    interaction:{{mode:'index',intersect:false}},
    plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:function(c){{return c.dataset.label+': '+(c.parsed.y!=null?c.parsed.y+'%':'—');}}}}}}  }},
    scales:{{y:{{display:false,min:-5,max:ym,ticks:{{callback:function(v){{return v<0?'':v+'%';}}}},title:{{display:false}}}},x:{{display:false,offset:true,ticks:{{maxRotation:45}},grid:{{display:false}}}}}}}};}}
  function optsToday(datasets){{
    var o=opts(datasets);
    var ym=datasets?yMax(datasets):110;
    o.scales.y={{display:true,min:0,max:ym,ticks:{{callback:function(v){{return v+'%';}},stepSize:10}},grid:{{color:'rgba(100,116,139,0.15)'}}}};
    o.scales.x={{display:true,offset:true,ticks:{{maxRotation:0,color:'#475569',font:{{size:11}}}},grid:{{display:false}}}};
    o.layout={{padding:{{left:4,right:12,top:4,bottom:4}}}};
    return o;
  }}
  function leg(id,ds2){{
    var el=document.getElementById(id);if(!el)return;
    el.innerHTML=ds2.map(function(d,i){{
      var col=COLS[i]||d.borderColor||'#888';
      return '<span class="leg-item"><span class="leg-dot" style="background:'+col+'"></span>'+d.label+'</span>';
    }}).join('');
  }}

  var _selDay=null;  // обрана дата (ISO); null = today
  function initToday(targetDate){{
    var _now=new Date();
    var tod=localISO(_now);
    var dayISO=targetDate||_selDay||tod;
    _selDay=dayISO;
    var isToday=(dayISO===tod);
    var hd=HDATA[dayISO]||{{}};
    var hs=new Set();
    Object.values(hd).forEach(function(m){{Object.keys(m).forEach(function(h){{hs.add(parseInt(h));}});}});
    // Принцип: точка "HH:00" відображає ефективність попередньої години (H-1 … H).
    // Для сьогодні — поточна (неповна) година _rh представлена окремою точкою REPORT_HM.
    // Для минулих днів — всі години вважаються завершеними.
    var _rh = isToday ? (REPORT_HM ? parseInt(REPORT_HM.split(':')[0],10) : _now.getHours()) : 24;
    var hours=Array.from(hs).filter(function(h){{return h<_rh;}}).sort(function(a,b){{return a-b;}});
    var labels=hours.map(function(h){{var e=(h+1)%24;return (e<10?'0':'')+e+':00';}});
    if(isToday&&REPORT_HM) labels.push(REPORT_HM);
    var d2=ds(labels,function(k,lbl){{
      var h;
      if(isToday&&lbl===REPORT_HM){{ h=_rh; }}
      else {{ h=parseInt(lbl,10)-1; if(h<0) h=23; }}
      return (hd[k]&&hd[k][String(h)]!=null)?hd[k][String(h)]:0;
    }});
    if(tCh)tCh.destroy();
    var ctx=document.getElementById('effChartToday');if(!ctx)return;
    var _plugins=isToday?[reportLinePlugin]:[];
    tCh=new Chart(ctx.getContext('2d'),{{type:'line',data:{{labels:labels,datasets:d2}},options:optsToday(d2),plugins:_plugins}});
    requestAnimationFrame(function(){{tCh.resize();var sc=ctx.closest('.chart-scroll-outer');if(sc)sc.scrollLeft=sc.scrollWidth;}});
    leg('legend-today',d2);
    var tt=document.getElementById('today-chart-title');
    if(tt){{
      var dStr=dayISO.slice(8,10)+'.'+dayISO.slice(5,7)+'.'+dayISO.slice(0,4);
      tt.textContent=(isToday?'Today':dStr)+' — Hourly Efficiency';
    }}
    var ttb=document.getElementById('today-table-wrap');
    if(ttb){{
      var sel=getSelected();
      var selM=sel.filter(function(k){{return k!=='SITE';}});
      var mAvgs={{}};
      selM.forEach(function(k){{
        if(isToday) mAvgs[k]=(TEFF&&TEFF[k]!=null)?TEFF[k]:0;
        else        mAvgs[k]=(ALL[dayISO]&&ALL[dayISO][k]!=null)?ALL[dayISO][k]:0;
      }});
      var grandAvg=selM.length?Math.round(selM.reduce(function(s,k){{return s+mAvgs[k];}},0)/selM.length):0;
      var dayStr=dayISO.slice(8,10)+'.'+dayISO.slice(5,7)+'.'+dayISO.slice(0,4);
      var hdr=sel.map(function(k){{var si=MK.indexOf(k);return '<th>'+SK[si]+'</th>';}}).join('');
      var cells=sel.map(function(k){{
        var avg=k==='SITE'?grandAvg:(mAvgs[k]||0);
        var col=avg>=75?'#22c55e':avg>=50?'#f59e0b':'#ef4444';
        return '<td><b style="color:'+col+'">'+avg+'%</b></td>';
      }}).join('');
      ttb.innerHTML=
        '<table class="stats-table"><thead><tr><th>Date</th>'+hdr+'</tr></thead>'+
        '<tbody><tr><td><b>'+dayStr+'</b></td>'+cells+'</tr></tbody></table>';
    }}
  }}

  // Селектор днів над Today-графіком: вантажимо ISO-дати з HDATA за останні 7 днів.
  function _initDaySelector(){{
    var el=document.getElementById('today-day-selector');
    if(!el) return;
    el.innerHTML='';
    var todISO=localISO(new Date());
    // Показуємо останні 7 днів завжди (навіть якщо в HDATA нема даних для якогось дня —
    // кнопка буде, чарт покаже порожньо).
    var days=[];
    for(var i=0;i<7;i++){{
      var dt=new Date();dt.setDate(dt.getDate()-i);
      days.push(localISO(dt));
    }}
    days.forEach(function(d){{
      var btn=document.createElement('button');
      btn.type='button';
      var isTod=(d===todISO);
      btn.textContent=isTod?'Today':(d.slice(8,10)+'.'+d.slice(5,7));
      btn.dataset.date=d;
      btn.style.cssText='cursor:pointer;font-size:0.75rem;font-weight:700;padding:4px 10px;'
        +'border-radius:20px;border:2px solid #cbd5e1;background:#fff;color:#1e293b;transition:background .15s';
      btn.addEventListener('click',function(){{
        _selDay=btn.dataset.date;
        Array.prototype.forEach.call(el.querySelectorAll('button'),function(b){{
          var on=(b.dataset.date===_selDay);
          b.style.background=on?'#3b82f6':'#fff';
          b.style.color=on?'#fff':'#1e293b';
          b.style.borderColor=on?'#3b82f6':'#cbd5e1';
        }});
        initToday(_selDay);
      }});
      el.appendChild(btn);
    }});
    // Вибираємо today за замовчуванням
    var todBtn=el.querySelector('button[data-date="'+todISO+'"]');
    if(todBtn){{
      todBtn.style.background='#3b82f6';
      todBtn.style.color='#fff';
      todBtn.style.borderColor='#3b82f6';
    }}
  }}

  function parseLblDate(lbl){{
    var p=lbl.split('.');return new Date(Number(p[2]),Number(p[1])-1,Number(p[0]));
  }}
  function getISOWeek(d){{
    var tmp=new Date(d.getFullYear(),d.getMonth(),d.getDate());
    var dayNum=tmp.getDay()||7;tmp.setDate(tmp.getDate()+4-dayNum);
    var y0=new Date(tmp.getFullYear(),0,1);
    return Math.ceil((((tmp-y0)/86400000)+1)/7);
  }}
  var weekendPeriodPlugin={{
    id:'weekendPeriod',
    afterLayout(chart){{
      var ca=chart.chartArea;
      ca.left=0;ca.right=chart.width;ca.width=chart.width;
      ca.top=0;ca.bottom=chart.height;ca.height=chart.height;
      var xS=chart.scales.x;
      if(xS){{xS.left=0;xS.right=chart.width;xS.width=chart.width;xS._startPixel=0;xS._endPixel=chart.width;xS._length=chart.width;}}
      var yS=chart.scales.y;
      if(yS){{yS.top=0;yS.bottom=chart.height;yS.height=chart.height;yS._startPixel=0;yS._endPixel=chart.height;yS._length=chart.height;}}
    }},
    beforeDraw(chart){{
      var c=chart.chartArea,xScale=chart.scales.x;
      if(!c) return;
      var ctx2=chart.ctx;
      ctx2.save();ctx2.fillStyle='#f8fafc';ctx2.fillRect(0,0,chart.width,chart.height);ctx2.restore();
      // Horizontal percentage grid lines (10%…100%)
      var yS2=chart.scales.y;
      [0,10,20,30,40,50,60,70,80,90,100].forEach(function(pct){{
        var yPx=yS2.getPixelForValue(pct);
        ctx2.save();
        ctx2.strokeStyle='rgba(100,116,139,0.18)';ctx2.lineWidth=1;ctx2.setLineDash([]);
        ctx2.beginPath();ctx2.moveTo(0,yPx);ctx2.lineTo(chart.width,yPx);ctx2.stroke();
        ctx2.fillStyle='rgba(100,116,139,0.5)';ctx2.font='9px sans-serif';ctx2.textAlign='left';
        ctx2.fillText(pct+'%',3,yPx-2);
        ctx2.restore();
      }});
      var labels=chart.data.labels,n=labels.length;if(n<2)return;
      var step=xScale.getPixelForValue(1)-xScale.getPixelForValue(0);
      // Weekend shading + SAT/SUN label + day separator lines + date labels
      var dayStep=step>=28?1:Math.ceil(28/step);
      labels.forEach(function(lbl,i){{
        var dow=parseLblDate(lbl).getDay();
        var cx=xScale.getPixelForValue(i);
        if(dow===0||dow===6){{
          ctx2.save();
          ctx2.fillStyle='rgba(255,180,0,0.18)';
          ctx2.fillRect(cx-step/2,0,step,chart.height);
          ctx2.fillStyle='rgba(180,100,0,0.6)';
          ctx2.font='bold 10px sans-serif';ctx2.textAlign='center';
          ctx2.fillText(dow===6?'SAT':'SUN',cx,c.top+11);
          ctx2.restore();
        }}else if(i%dayStep===0){{
          var dn=['SUN','MON','TUE','WED','THU','FRI','SAT'][dow];
          ctx2.save();
          ctx2.fillStyle='rgba(100,116,139,0.55)';
          ctx2.font='bold 10px sans-serif';ctx2.textAlign='center';
          ctx2.fillText(dn,cx,c.top+11);
          ctx2.restore();
        }}
        // Day separator line at left edge of each column (between prev day and this day)
        if(i>0){{
          var sepX=cx-step/2;
          ctx2.save();
          ctx2.strokeStyle='#475569'; ctx2.lineWidth=1;
          ctx2.setLineDash([]);
          ctx2.beginPath();ctx2.moveTo(sepX,0);ctx2.lineTo(sepX,chart.height);ctx2.stroke();
          ctx2.restore();
        }}
      }});
    }},
    afterDraw(chart){{
      var xScale=chart.scales.x;
      var labels=chart.data.labels;
      if(!labels||labels.length<2)return;
      var ctx2=chart.ctx;
      var step=xScale.getPixelForValue(1)-xScale.getPixelForValue(0);
      // Week number labels + dashed separators
      var wg={{}};
      labels.forEach(function(lbl,i){{
        var wn=getISOWeek(parseLblDate(lbl));
        if(!wg[wn])wg[wn]=[];wg[wn].push(i);
      }});
      var weeks=Object.entries(wg).sort(function(a,b){{return Number(a[0])-Number(b[0]);}});
      weeks.forEach(function(entry,gi){{
        var wn=entry[0],idxs=entry[1];
        var avgX=idxs.reduce(function(s,i){{return s+xScale.getPixelForValue(i);}},0)/idxs.length;
        ctx2.save();
        ctx2.fillStyle='rgba(248,250,252,0.8)';
        ctx2.fillRect(avgX-14,chart.height-12,28,12);
        ctx2.fillStyle='rgba(80,80,180,0.85)';
        ctx2.font='bold 10px sans-serif';ctx2.textAlign='center';
        ctx2.fillText('W'+wn,avgX,chart.height-3);
        if(gi>0){{
          var prev=weeks[gi-1][1];
          var sepX=(xScale.getPixelForValue(idxs[0])+xScale.getPixelForValue(prev[prev.length-1]))/2;
          ctx2.strokeStyle='rgba(80,80,180,0.55)';ctx2.lineWidth=3;
          ctx2.setLineDash([6,3]);ctx2.beginPath();
          ctx2.moveTo(sepX,0);ctx2.lineTo(sepX,chart.height);ctx2.stroke();
        }}
        ctx2.restore();
      }});
    }}
  }};
  function euToISO(s){{var p=s.split('.');return p.length===3?p[2]+'-'+p[1]+'-'+p[0]:s;}}
  function isoToEu(s){{return s.slice(8,10)+'.'+s.slice(5,7)+'.'+s.slice(0,4);}}
  function updatePeriodChart(){{
    var from=euToISO(document.getElementById('stat-from').value);
    var to=euToISO(document.getElementById('stat-to').value);
    // Generate ALL dates in range so both Period Trend and Batch Timeline
    // have identical column counts — required for x-axis alignment
    var dates=[];
    {{var _c=new Date(from+'T00:00:00'),_e=new Date(to+'T00:00:00');
      while(_c<=_e){{dates.push(localISO(_c));_c.setDate(_c.getDate()+1);}}}}
    function fmtDate(s){{return s.slice(8,10)+'.'+s.slice(5,7)+'.'+s.slice(0,4);}}
    // Принцип: точка з датою D відображає ефективність попереднього дня (інтервал D-1 00:00 … D 00:00).
    function _prevDay(iso){{var dt=new Date(iso+'T00:00:00');dt.setDate(dt.getDate()-1);return localISO(dt);}}
    var d2=ds(dates,function(k,d){{var pd=_prevDay(d);return ALL[pd]&&ALL[pd][k]!=null?ALL[pd][k]:0;}});
    // Weekly average curve
    (function(){{
      var sel=getSelected();
      var selM=sel.filter(function(k){{return k!=='SITE';}});
      if(!selM.length)return;
      // Group dates by ISO week (year+week key to handle year boundaries).
      // Агрегація проводиться за повним тижнем (дані дня D), а лише ПОЗИЦІЯ точки
      // зсувається на кінець тижня відповідно до принципу "кінець періоду".
      var weekGroups={{}};
      dates.forEach(function(d){{
        var dt=new Date(d+'T00:00:00');
        var wn=getISOWeek(dt);
        var yr=dt.getFullYear();
        if(wn>=52&&dt.getMonth()===0)yr--;
        if(wn===1&&dt.getMonth()===11)yr++;
        var wk=yr+'_'+wn;
        if(!weekGroups[wk])weekGroups[wk]=[];
        weekGroups[wk].push(d);
      }});
      // Per-week average: mean of daily averages (each day first averaged across machines)
      var weekAvg={{}};
      Object.keys(weekGroups).forEach(function(wk){{
        var dayAvgs=[];
        weekGroups[wk].forEach(function(d){{
          var vals=selM.map(function(k){{return (ALL[d]&&ALL[d][k]!=null)?ALL[d][k]:0;}});
          var dayAvg=vals.reduce(function(a,b){{return a+b;}},0)/vals.length;
          // Include only days when machines actually worked (weekdays always, weekends only if dayAvg > 0)
          var dow=new Date(d+'T00:00:00').getDay();
          if((dow===0||dow===6)&&dayAvg===0) return;
          dayAvgs.push(dayAvg);
        }});
        // Always divide by 5 (standard work week): weekday work is expected,
        // weekend work adds bonus without inflating the denominator
        weekAvg[wk]=Math.min(100,Math.round(dayAvgs.reduce(function(a,b){{return a+b;}},0)/5));
      }});
      // Принцип "кінець періоду": ставимо середньотижневу точку на останню дату групи
      // (межа "кінець тижня" в шкалі labels).
      var weekCenter={{}};
      Object.keys(weekGroups).forEach(function(wk){{
        var wdates=weekGroups[wk];
        var end=wdates[wdates.length-1];
        weekCenter[end]=weekAvg[wk];
      }});
      var weekData=dates.map(function(d){{
        return (weekCenter[d]!=null)?weekCenter[d]:null;
      }});
      d2.push({{
        label:'Week avg',data:weekData,
        borderColor:'rgba(30,41,59,0.85)',backgroundColor:'rgba(30,41,59,0.06)',
        tension:0.4,pointRadius:5,pointHoverRadius:7,
        pointBackgroundColor:'rgba(30,41,59,0.85)',
        borderWidth:2.5,borderDash:[6,3],
        stepped:false,fill:false,spanGaps:true,order:10
      }});
    }})();
    if(pCh)pCh.destroy();
    var ctx=document.getElementById('effChartPeriod');if(!ctx)return;
    var _wrap=ctx.closest('.chart-wrap');
    var _minW=Math.max(600,dates.length*22);
    if(_wrap){{_wrap.style.width='';_wrap.style.minWidth=_minW+'px';}}
    pCh=new Chart(ctx.getContext('2d'),{{type:'line',plugins:[weekendPeriodPlugin],data:{{labels:dates.map(fmtDate),datasets:d2}},options:opts(d2)}});
    requestAnimationFrame(function(){{
      var _actualW=(_wrap&&_wrap.offsetWidth>0)?_wrap.offsetWidth:_minW;
      if(_wrap)_wrap.style.width=_actualW+'px';
      if(pCh)pCh.resize(_actualW,350);
      if(window.updateBatchGantt)window.updateBatchGantt(from,to,_actualW);
      // Auto-scroll both panels to end when content overflows
      var _pSc=ctx.closest('.chart-scroll-outer');
      var _gSc=document.getElementById('gantt-scroll-outer');
      if(_pSc&&_pSc.scrollWidth>_pSc.clientWidth){{_pSc.scrollLeft=_pSc.scrollWidth;}}
      if(_gSc&&_gSc.scrollWidth>_gSc.clientWidth){{_gSc.scrollLeft=_gSc.scrollWidth;}}
    }});
    leg('legend-period',d2);
  }}

  window.initToday=initToday;
  window.initDaySelector=_initDaySelector;
  window.updatePeriodChart=updatePeriodChart;
  window.setRange=function(n){{
    var to=new Date(),from=new Date();from.setDate(to.getDate()-n+1);
    document.getElementById('stat-from').value=isoToEu(localISO(from));
    document.getElementById('stat-to').value=isoToEu(localISO(to));
    updatePeriodChart();
  }};
  // Scroll-to-top
  var stBtn=document.getElementById('scroll-top');
  if(stBtn) window.addEventListener('scroll',function(){{stBtn.classList.toggle('visible',window.scrollY>400);}});
}})();

// ── Batch Gantt ──────────────────────────────────────────────────
(function(){{
  var RAW_ALL={_cdata_js};
  // Розпаковуємо стислі поля: дата "260419"→"2026-04-19", час "0830"→"08:30",
  // програма "330-903B" → "WF330-903B" (без суфіксу .MIN)
  RAW_ALL.forEach(function(c){{
    if(c.d&&c.d.length===6) c.d='20'+c.d.substr(0,2)+'-'+c.d.substr(2,2)+'-'+c.d.substr(4,2);
    if(c.s&&c.s.length===4) c.s=c.s.substr(0,2)+':'+c.s.substr(2,2);
    if(c.e&&c.e.length===4) c.e=c.e.substr(0,2)+':'+c.e.substr(2,2);
    if(c.p) c.p='WF'+c.p;
  }});
  var progList=[],progColor={{}};
  var PALETTE=['#3b82f6','#22c55e','#f59e0b','#0ea5e9','#a855f7','#06b6d4','#f97316','#65a30d','#84cc16','#14b8a6'];
  RAW_ALL.forEach(function(c){{
    if(progList.indexOf(c.p)===-1){{progList.push(c.p);progColor[c.p]=PALETTE[(progList.length-1)%PALETTE.length];}}
  }});

  var ROW_H=32, PAD=4, LABEL_W=0, TICK_H=22;
  var PX_PER_DAY=28;
  var cv=document.getElementById('batchGantt');
  var tip=document.getElementById('batch-tooltip');
  var barMeta=[], gapMeta=[], CDATA=[], machines=[], dates=[];

  function buildFromRange(from, to, forcedW){{
    // Filter RAW_ALL by date range
    var raw=RAW_ALL.filter(function(c){{return (!from||c.d>=from)&&(!to||c.d<=to);}});
    // Session-based grouping: merge consecutive cycles of the same program
    // into one block if the gap between end of previous and start of next is ≤SESSION_GAP min.
    // Cycles with gaps > SESSION_GAP become separate blocks (separate production runs).
    var SESSION_GAP=60;
    var sorted=raw.slice().sort(function(a,b){{
      if(a.d!==b.d) return a.d<b.d?-1:1;
      if(a.m!==b.m) return a.m<b.m?-1:1;
      return (a.s||'')<(b.s||'')?-1:1;
    }});
    CDATA=[];
    var sessMap={{}};
    sorted.forEach(function(c){{
      var key=c.m+'|'+c.d+'|'+c.p;
      var sess=sessMap[key];
      // Cross-midnight merge: if no session found for this date, check previous day
      if(!sess&&c.s){{
        var _pd=new Date(c.d+'T00:00:00');_pd.setDate(_pd.getDate()-1);
        var _prevKey=c.m+'|'+localISO(_pd)+'|'+c.p;
        var _prevSess=sessMap[_prevKey];
        if(_prevSess){{
          var _sh=parseInt(c.s.split(':')[0]),_sm=parseInt(c.s.split(':')[1]||0);
          var _startMin=_sh*60+_sm;
          if(!_prevSess.e){{
            // Previous session was still ongoing at midnight (end_time="—") — this IS the cross-midnight case
            if(_startMin<=SESSION_GAP){{sess=_prevSess;sessMap[key]=_prevSess;}}
          }} else {{
            var _eh=parseInt(_prevSess.e.split(':')[0]),_em=parseInt(_prevSess.e.split(':')[1]||0);
            var _gap=_startMin+1440-(_eh*60+_em);
            if(_gap<=SESSION_GAP){{sess=_prevSess;sessMap[key]=_prevSess;}}
          }}
        }}
      }}
      if(sess&&c.s){{
        var canExtend=false;
        if(!sess.e){{
          // ongoing session — allow extending only if start is not too far ahead
          var sh2=parseInt(c.s.split(':')[0]),sm2=parseInt(c.s.split(':')[1]||0);
          var sh0=parseInt(sess.s.split(':')[0]),sm0=parseInt(sess.s.split(':')[1]||0);
          canExtend=((sh2*60+sm2)-(sh0*60+sm0)<=SESSION_GAP*4);
        }}
        else{{
          var eh2=parseInt(sess.e.split(':')[0]),em2=parseInt(sess.e.split(':')[1]||0);
          var sh2=parseInt(c.s.split(':')[0]),sm2=parseInt(c.s.split(':')[1]||0);
          var gap=(sh2*60+sm2)-(eh2*60+em2);
          // allow small overlap (gap<0) as well as normal gap
          canExtend=(gap<=SESSION_GAP);
        }}
        if(canExtend){{
          if(c.e&&(!sess.e_d||c.d>sess.e_d||(c.d===sess.e_d&&c.e>sess.e))){{sess.e=c.e;sess.e_d=c.d;}}
          sess.dur=(sess.dur||0)+(c.dur||0);
          return;
        }}
      }}
      var ns={{d:c.d,m:c.m,p:c.p,s:c.s,e:c.e,dur:c.dur||0,e_d:c.d}};
      CDATA.push(ns);
      sessMap[key]=ns;
    }});
    CDATA.sort(function(a,b){{return a.d<b.d?-1:a.d>b.d?1:0;}});
    machines=[];
    CDATA.forEach(function(c){{
      if(machines.indexOf(c.m)===-1)machines.push(c.m);
    }});
    machines.sort();
    // Generate ALL dates in range (not just dates with data) so columns
    // match the Period Trend chart which shows every date in the range
    dates=[];
    if(from&&to){{
      var cur=new Date(from+'T00:00:00');
      var end=new Date(to+'T00:00:00');
      while(cur<=end){{
        dates.push(localISO(cur));
        cur.setDate(cur.getDate()+1);
      }}
    }} else {{
      var dateSet={{}};
      CDATA.forEach(function(c){{dateSet[c.d]=1;}});
      dates=Object.keys(dateSet).sort();
    }}
    // Compute PX_PER_DAY from Period Trend chart-wrap width (forcedW passed directly avoids offsetWidth timing issues)
    if(dates.length){{
      var _availW=forcedW||0;
      if(!_availW){{var _chartEl=document.getElementById('effChartPeriod');var _cw=_chartEl?_chartEl.closest('.chart-wrap'):null;_availW=(_cw&&_cw.offsetWidth>0)?_cw.offsetWidth:0;}}
      if(!_availW)_availW=Math.max(600,dates.length*22+60);
      PX_PER_DAY=_availW/Math.max(dates.length,1);
    }}
  }}

  function drawGantt(forcedW){{
    if(!dates.length){{ cv.style.width='0'; cv.style.height='0'; return; }}
    var canvasW=forcedW||0;
    if(!canvasW){{var _chartEl=document.getElementById('effChartPeriod');var _cw2=_chartEl?_chartEl.closest('.chart-wrap'):null;canvasW=(_cw2&&_cw2.offsetWidth>0)?_cw2.offsetWidth:0;}}
    if(!canvasW)canvasW=Math.max(600,dates.length*22+60);
    // Sync Gantt chart-wrap width to match Period Trend exactly
    var _gcw=cv.closest('.chart-wrap');
    if(_gcw){{_gcw.style.minWidth=canvasW+'px';_gcw.style.width=canvasW+'px';}}
    var canvasH=350;
    if(machines.length>0){{ROW_H=Math.max(14,Math.floor((canvasH-TICK_H-PAD*(machines.length+1))/machines.length));}}
    var dpr=window.devicePixelRatio||1;
    cv.style.width=canvasW+'px'; cv.style.height=canvasH+'px';
    cv.width=Math.round(canvasW*dpr); cv.height=Math.round(canvasH*dpr);
    var ctx=cv.getContext('2d'); ctx.scale(dpr,dpr);
    barMeta=[];

    ctx.fillStyle='#f8fafc'; ctx.fillRect(0,0,canvasW,canvasH);

    dates.forEach(function(d,i){{
      var x=LABEL_W+i*PX_PER_DAY;
      var dow=new Date(d+'T00:00:00').getDay();
      if(dow===0||dow===6){{
        ctx.fillStyle='rgba(255,180,0,0.18)';
        ctx.fillRect(x,0,PX_PER_DAY,canvasH-TICK_H);
        ctx.save();ctx.fillStyle='rgba(180,100,0,0.6)';ctx.font='bold 10px sans-serif';ctx.textAlign='center';
        ctx.fillText(dow===6?'SAT':'SUN',x+PX_PER_DAY/2,11);ctx.restore();
      }} else {{
        // Non-working hours: dark gray fill
        // Mon(1): dark 0-7h, 19-24h
        // Tue(2): dark 0-6.5h (Mon ends 19:00, no carry-over)
        // Wed(3): dark 0.5-6.5h (Tue ends 00:30 Wed, so 00:00-00:30 is working)
        // Thu(4): dark 0.5-6.5h (Wed ends 00:30 Thu)
        // Fri(5): dark 0.5-7h, 19-24h (Thu ends 00:30 Fri)
        ctx.fillStyle='rgba(71,85,105,0.45)';
        if(dow===1){{
          ctx.fillRect(x,0,Math.round(7/24*PX_PER_DAY),canvasH-TICK_H);
          ctx.fillRect(x+Math.round(19/24*PX_PER_DAY),0,PX_PER_DAY-Math.round(19/24*PX_PER_DAY),canvasH-TICK_H);
        }} else if(dow===5){{
          ctx.fillRect(x+Math.round(0.5/24*PX_PER_DAY),0,Math.round(7/24*PX_PER_DAY)-Math.round(0.5/24*PX_PER_DAY),canvasH-TICK_H);
          ctx.fillRect(x+Math.round(19/24*PX_PER_DAY),0,PX_PER_DAY-Math.round(19/24*PX_PER_DAY),canvasH-TICK_H);
        }} else if(dow===2){{
          ctx.fillRect(x,0,Math.round(6.5/24*PX_PER_DAY),canvasH-TICK_H);
        }} else {{
          // Wed(3), Thu(4): 00:30-06:30 is non-working; 00:00-00:30 is working carry-over
          ctx.fillRect(x+Math.round(0.5/24*PX_PER_DAY),0,Math.round(6.5/24*PX_PER_DAY)-Math.round(0.5/24*PX_PER_DAY),canvasH-TICK_H);
        }}
        // Working hours: white background (program bars will cover it)
        ctx.fillStyle='rgba(255,255,255,1)';
        if(dow===1){{
          ctx.fillRect(x+Math.round(7/24*PX_PER_DAY),0,Math.round(19/24*PX_PER_DAY)-Math.round(7/24*PX_PER_DAY),canvasH-TICK_H);
        }} else if(dow===5){{
          ctx.fillRect(x,0,Math.round(0.5/24*PX_PER_DAY),canvasH-TICK_H);
          ctx.fillRect(x+Math.round(7/24*PX_PER_DAY),0,Math.round(19/24*PX_PER_DAY)-Math.round(7/24*PX_PER_DAY),canvasH-TICK_H);
        }} else if(dow===2){{
          ctx.fillRect(x+Math.round(6.5/24*PX_PER_DAY),0,PX_PER_DAY-Math.round(6.5/24*PX_PER_DAY),canvasH-TICK_H);
        }} else {{
          // Wed(3), Thu(4): 00:00-00:30 working + 06:30-24:00 working
          ctx.fillRect(x,0,Math.round(0.5/24*PX_PER_DAY),canvasH-TICK_H);
          ctx.fillRect(x+Math.round(6.5/24*PX_PER_DAY),0,PX_PER_DAY-Math.round(6.5/24*PX_PER_DAY),canvasH-TICK_H);
        }}
      }}
      ctx.strokeStyle='#475569'; ctx.lineWidth=1; ctx.setLineDash([]);
      ctx.beginPath(); ctx.moveTo(x+0.5,0); ctx.lineTo(x+0.5,canvasH-TICK_H); ctx.stroke();
      var step=PX_PER_DAY>=28?1:Math.ceil(28/PX_PER_DAY);
      if(i%step===0){{
        ctx.fillStyle='#64748b'; ctx.font='10px sans-serif'; ctx.textAlign='center';
        ctx.fillText(d.slice(8)+'.'+d.slice(5,7), x+PX_PER_DAY/2, canvasH-6);
      }}
    }});

    // ── Machine separator lines ─────────────────────────────────────
    machines.forEach(function(m,ri){{
      var y=PAD+ri*(ROW_H+PAD);
      ctx.strokeStyle='#475569'; ctx.lineWidth=1; ctx.setLineDash([]);
      ctx.beginPath(); ctx.moveTo(0,y+ROW_H+PAD/2); ctx.lineTo(canvasW,y+ROW_H+PAD/2); ctx.stroke();
    }});

    // ── Collect gap metadata (idle between programs, same machine+day) ─
    var byMD={{}};
    CDATA.forEach(function(c){{
      var k=c.m+'|'+c.d;
      if(!byMD[k])byMD[k]=[];
      byMD[k].push(c);
    }});
    gapMeta=[];
    Object.keys(byMD).forEach(function(k){{
      var items=byMD[k].slice().sort(function(a,b){{return (a.s||'')<(b.s||'')?-1:1;}});
      var parts=k.split('|'), m=parts[0], d=parts[1];
      var ri=machines.indexOf(m), di=dates.indexOf(d);
      if(ri===-1||di===-1) return;
      var x=LABEL_W+di*PX_PER_DAY;
      var y=PAD+ri*(ROW_H+PAD);
      for(var i=0;i<items.length-1;i++){{
        var cur=items[i],nxt=items[i+1];
        if(!cur.e||!nxt.s) continue;
        var ceh=parseInt(cur.e.split(':')[0]),cem=parseInt(cur.e.split(':')[1]||0);
        var nsh=parseInt(nxt.s.split(':')[0]),nsm=parseInt(nxt.s.split(':')[1]||0);
        var gapMin=(nsh*60+nsm)-(ceh*60+cem);
        if(gapMin<15) continue;
        var gx=x+Math.round((ceh*60+cem)/1440*PX_PER_DAY);
        var gex=x+Math.round((nsh*60+nsm)/1440*PX_PER_DAY);
        var gw=Math.max(gex-gx,2);
        gapMeta.push({{x:gx,y:y+3,w:gw,h:ROW_H-6,gapMin:gapMin,m:m,d:d,s:cur.e,e:nxt.s}});
      }}
    }});

    // ── Program bars (multi-day aware) ────────────────────────────────
    CDATA.forEach(function(c){{
      var ri=machines.indexOf(c.m); if(ri===-1) return;
      var startDi=dates.indexOf(c.d); if(startDi===-1) return;
      if(!c.s) return;
      var sh=parseInt(c.s.split(':')[0]),sm=parseInt(c.s.split(':')[1]||0);
      var eh=c.e?parseInt(c.e.split(':')[0]):sh, em=c.e?parseInt(c.e.split(':')[1]||0):sm;
      var endDi=c.e_d?dates.indexOf(c.e_d):startDi;
      if(endDi===-1) endDi=startDi;
      var y=PAD+ri*(ROW_H+PAD);
      for(var dayIdx=startDi;dayIdx<=Math.min(endDi,dates.length-1);dayIdx++){{
        var x_d=LABEL_W+dayIdx*PX_PER_DAY;
        var segS=(dayIdx===startDi)?sh*60+sm:0;
        var segE=(dayIdx===endDi)?eh*60+em:1440;
        if(segS>=segE) continue;
        var bx=x_d+Math.round(segS/1440*PX_PER_DAY);
        var ex2=x_d+Math.round(segE/1440*PX_PER_DAY);
        var bw=Math.max(ex2-bx,4);
        ctx.fillStyle=progColor[c.p]||'#94a3b8';
        ctx.beginPath(); ctx.roundRect(bx,y+3,bw,ROW_H-6,3); ctx.fill();
        if(bw>36){{
          ctx.fillStyle='rgba(255,255,255,0.92)';
          ctx.font='bold 9px sans-serif'; ctx.textAlign='left';
          ctx.save(); ctx.beginPath(); ctx.rect(bx+2,y+3,bw-4,ROW_H-6); ctx.clip();
          ctx.fillText(c.p,bx+5,y+ROW_H/2+3);
          ctx.restore();
        }}
        barMeta.push({{x:bx,y:y+3,w:bw,h:ROW_H-6,c:c}});
      }}
    }});

    // Save clean canvas (without machine names) for sticky redraw on scroll
    _ganttClean=ctx.getImageData(0,0,cv.width,cv.height);
    _drawMachineNames();
  }}

  // ── Sticky machine labels (redrawn on scroll) ─────────────────────
  var _ganttClean=null;
  function _drawMachineNames(){{
    if(!_ganttClean||!machines.length)return;
    var _ctx=cv.getContext('2d');
    _ctx.putImageData(_ganttClean,0,0);
    var _gSc2=document.getElementById('gantt-scroll-outer');
    var _scrollX=_gSc2?_gSc2.scrollLeft:0;
    machines.forEach(function(m,ri){{
      var y=PAD+ri*(ROW_H+PAD);
      var shortM=m.indexOf('_')!==-1?m.split('_')[0]:m;
      _ctx.save();
      _ctx.font='bold 16px sans-serif';_ctx.textAlign='left';
      _ctx.shadowColor='rgba(0,0,0,0.8)';_ctx.shadowBlur=4;
      _ctx.lineWidth=3;_ctx.strokeStyle='rgba(0,0,0,0.6)';
      _ctx.strokeText(shortM,_scrollX+6,y+ROW_H/2+4);
      _ctx.fillStyle='#ffffff';_ctx.shadowBlur=0;
      _ctx.fillText(shortM,_scrollX+6,y+ROW_H/2+4);
      _ctx.restore();
    }});
  }}
  (function(){{
    var _gSc3=document.getElementById('gantt-scroll-outer');
    if(_gSc3){{var _raf;_gSc3.addEventListener('scroll',function(){{
      if(_raf)return;_raf=requestAnimationFrame(function(){{_drawMachineNames();_raf=null;}});
    }});}}
  }})();

  // Initial draw: last 7 days
  (function(){{
    var to=new Date(),from=new Date();from.setDate(to.getDate()-6);
    buildFromRange(localISO(from), localISO(to));
    drawGantt();
  }})();

  // Exposed so Period Trend controls can sync this chart
  window.updateBatchGantt=function(from,to,forcedW){{
    buildFromRange(from,to,forcedW);
    drawGantt(forcedW);
  }};

  // Tooltip
  cv.addEventListener('mousemove',function(e){{
    var rect=cv.getBoundingClientRect();
    var scaleX=parseFloat(cv.style.width)/rect.width;
    var mx=(e.clientX-rect.left)*scaleX, my=(e.clientY-rect.top)*scaleX;
    // Check program bars first
    var found=null;
    for(var i=barMeta.length-1;i>=0;i--){{
      var b=barMeta[i];
      if(mx>=b.x&&mx<=b.x+b.w&&my>=b.y&&my<=b.y+b.h){{found=b;break;}}
    }}
    function showTip(html, cx, cy){{
      tip.style.display='block';
      tip.innerHTML=html;
      var tw=tip.offsetWidth||200, th=tip.offsetHeight||70;
      var tx=cx+14, ty=cy-32;
      if(tx+tw>window.innerWidth-8) tx=cx-tw-14;
      if(tx<8) tx=8;
      if(ty+th>window.innerHeight-8) ty=cy-th-8;
      if(ty<8) ty=cy+14;
      tip.style.left=tx+'px'; tip.style.top=ty+'px';
    }}
    if(found){{
      var dm=found.c.dur?Math.round(found.c.dur):0;
      var durStr=dm>0?(dm>=60?(Math.floor(dm/60)+'h'+(dm%60?'\u00a0'+dm%60+'m':'')):dm+'\u00a0min'):'';
      showTip('<b>'+found.c.m+'</b> — '+found.c.p+'<br>'+found.c.d+'&nbsp;'+found.c.s+'–'+(found.c.e||'…')+(durStr?'<br>'+durStr:''),e.clientX,e.clientY);
      return;
    }}
    // Check gap bars
    var foundGap=null;
    for(var j=gapMeta.length-1;j>=0;j--){{
      var g=gapMeta[j];
      if(mx>=g.x&&mx<=g.x+g.w&&my>=g.y&&my<=g.y+g.h){{foundGap=g;break;}}
    }}
    if(foundGap){{
      var h=Math.floor(foundGap.gapMin/60),m2=foundGap.gapMin%60;
      var durStr=h>0?(h+'h '+(m2?m2+'m':'')):m2+'m';
      showTip('⏸ <b>Idle: '+durStr+'</b><br>'+foundGap.m+' &nbsp;'+foundGap.d+'<br>'+foundGap.s+' – '+foundGap.e,e.clientX,e.clientY);
    }} else tip.style.display='none';
  }});
  cv.addEventListener('mouseleave',function(){{tip.style.display='none';}});
}})();

// ── Nav drawer (mobile) ───────────────────────────────────────────
(function(){{
  var toggle=document.getElementById('nav-toggle');
  var sidebar=document.getElementById('nav-sidebar');
  var overlay=document.getElementById('nav-overlay');
  if(!toggle||!sidebar||!overlay) return;
  function openNav(){{sidebar.classList.add('open');overlay.classList.add('open');}}
  function closeNav(){{sidebar.classList.remove('open');overlay.classList.remove('open');}}
  toggle.addEventListener('click',function(){{
    sidebar.classList.contains('open')?closeNav():openNav();
  }});
  overlay.addEventListener('click',closeNav);
  sidebar.querySelectorAll('.nav-btn').forEach(function(btn){{
    btn.addEventListener('click',closeNav);
  }});
}})();

document.addEventListener('DOMContentLoaded',function(){{
  // Scroll sync: Period Trend ↔ Batch Gantt
  (function(){{
    var pEl=document.getElementById('effChartPeriod');
    var pSc=pEl?pEl.closest('.chart-scroll-outer'):null;
    var gSc=document.getElementById('gantt-scroll-outer');
    if(!pSc||!gSc) return;
    var _sy=false;
    pSc.addEventListener('scroll',function(){{if(_sy)return;_sy=true;gSc.scrollLeft=pSc.scrollLeft;_sy=false;}});
    gSc.addEventListener('scroll',function(){{if(_sy)return;_sy=true;pSc.scrollLeft=gSc.scrollLeft;_sy=false;}});
  }})();
  setTimeout(function(){{requestAnimationFrame(function(){{
    if(window.initDaySelector) window.initDaySelector();
    if(window.initToday) window.initToday();
    if(window.setRange)  window.setRange(7);
  }});}},80);
}});
</script>
</body>
</html>"""
# =============================================================================
def kill_old_instances():
    """Kill any other running instances of factory_monitor.py"""
    my_pid = os.getpid()
    try:
        # Get all python processes with their command lines
        result = subprocess.run(
            ['wmic', 'process', 'where',
             "name like '%python%'",
             'get', 'ProcessId,CommandLine', '/format:list'],
            capture_output=True, text=True, timeout=10
        )
        lines = result.stdout.strip().split('\n')

        current_cmd = None
        current_pid = None
        killed = 0

        for line in lines:
            line = line.strip()
            if line.startswith('CommandLine='):
                current_cmd = line[len('CommandLine='):]
            elif line.startswith('ProcessId='):
                try:
                    current_pid = int(line[len('ProcessId='):])
                except ValueError:
                    current_pid = None

                if current_cmd and current_pid and current_pid != my_pid:
                    if 'factory_monitor' in current_cmd.lower():
                        try:
                            subprocess.run(
                                ['taskkill', '/PID', str(current_pid), '/F'],
                                capture_output=True, timeout=5
                            )
                            log(f"Killed old instance PID {current_pid}")
                            killed += 1
                        except Exception as e:
                            log(f"Failed to kill PID {current_pid}: {e}")

                current_cmd = None
                current_pid = None

        if killed:
            log(f"Killed {killed} old instance(s)")
        else:
            log("No old instances found")
    except Exception as e:
        log(f"kill_old_instances error: {e} — continuing anyway")


def main():
    import traceback as _tb
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    kill_old_instances()

    log("=" * 60)
    log("FACTORY MONITOR START — V14")
    log("=" * 60)

    # Step 2 — fetch data via WebAPI
    rows, mr_data = fetch_from_api()
    if not rows:
        log("No data received from API — aborting.")
        sys.exit(1)

    log(f"Rows loaded: {len(rows)}")
    log(f"Machines: {sorted(set(r['MachineName'] for r in rows if r.get('MachineName')))}")

    filtered, period_from, period_to = filter_last_hours(rows, HOURS_BACK)
    date_str = period_to.strftime("%Y-%m-%d")
    log(f"Period: {period_from.strftime('%H:%M')} – {period_to.strftime('%H:%M')} ({len(filtered)} rows)")

    # Step 3 — analyze
    try:
        log("── Step 3: Analyzing cycles ──")
        cycles = analyze_cycles(filtered)
        log(f"  Cycles: {sum(len(v) for v in cycles.values())}")

        log("── Step 3.1: Counter markers ──")
        counter_markers = get_counter_markers(mr_data, cycles)
        counter_machines = set(counter_markers.keys())
        cycles = split_cycles_by_counter(cycles, counter_markers)
        cycles, counter_markers = apply_start_to_start_cycles(cycles, counter_markers, mr_data)
        counter_markers = add_runstate_boundary_markers(counter_markers, filtered, counter_machines)
        log(f"  Counter machines: {sorted(counter_machines)}")

        log("── Step 3.2: Downtime ──")
        downtimes = analyze_downtime(filtered)

        log("── Step 3.3: Timeline ──")
        timeline_data = build_timeline_data(filtered, period_from, period_to)
        timeline_data = split_timeline_by_counter(timeline_data, counter_markers, period_from, period_to)
        log("  Timeline done")
    except Exception as e:
        log(f"✗ Analysis error: {e}")
        log(_tb.format_exc())
        sys.exit(1)

    try:
        conn = init_db()
        save_to_db(conn, date_str, cycles, downtimes)
        log("History saved to DB")
    except Exception as e:
        log(f"✗ DB error: {e}")
        log(_tb.format_exc())
        sys.exit(1)

    # Step 3.5 — load Excel target times
    log("── Step 3.5: Loading Excel target times ──")
    excel_targets = load_target_times()

    # Step 4 — report
    log("── Step 4: Generating report ──")
    try:
        html = generate_html(cycles, downtimes, period_from, period_to, timeline_data, conn, excel_targets, counter_markers)
    except Exception as e:
        log(f"✗ Error generating HTML: {e}")
        log(_tb.format_exc())
        raise

    conn.close()

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    log(f"Report saved: {OUTPUT_HTML}")

    # Step 5 — publish to GitHub Pages
    log("── Step 5: Publishing to GitHub Pages ──")
    publish_to_github(html)

    # Step 6 — Telegram alert (раз на годину, контролюється маркер-файлом)
    log("── Step 6: Telegram alert check ──")
    check_and_alert(downtimes, period_to, cycles, excel_targets)

    log("=" * 60)
    log("FACTORY MONITOR COMPLETE")
    log("=" * 60)

if __name__ == "__main__":
    main()